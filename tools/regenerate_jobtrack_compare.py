from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from urllib.parse import quote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GENERATED = ROOT / "reports/jobtrack_0610_compare/订单整理结果_源表规则修正.xlsx"
DEFAULT_REFERENCE = ROOT / "data/reference/2026 Job Track 0610.xlsx"
DEFAULT_AUDIT = ROOT / "reports/jobtrack_0610_compare/audit_源表规则修正.csv"
DEFAULT_DIFFS = ROOT / "reports/jobtrack_0610_compare/diffs_source_rules.csv"
DEFAULT_OVERSIZE = ROOT / "reports/jobtrack_0610_compare/oversize_source_rules_analysis.csv"
DEFAULT_HTML = ROOT / "outputs/problem_orders_tool/problem_orders.html"
DEFAULT_REPORT_HTML = ROOT / "reports/problem_orders_tool/problem_orders.html"
DEFAULT_SOURCE_DIR = ROOT / "data/input/order_excels_dedup"
ZERO_DECIMAL_DISPLAY_COLUMNS = {18, 19, 20, 21, 22, 23, 24}
EXPLAINED_BUILDER_PAIRS = {
    ("M & B", "M&B"),
    ("CARNARVON", "Carnarvon Timber & Hardware"),
    ("Coastview", "Coastview Australia Pty Ltd (River Stone Design)"),
    ("RIVERSTONE", "Coastview Australia Pty Ltd (River Stone Design)"),
    ("WB", "Webb & Brown-Neaves"),
    ("WEBB", "Webb & Brown-Neaves"),
    ("PLANT", "PLANET"),
}


def display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(Decimal(str(value)).normalize(), "f")
    return str(value).strip()


def normalized(value: object, col: int) -> str:
    text = display_value(value)
    if col == 2:
        text = " ".join(text.split())
        return str(int(text)) if text.isdigit() else text
    if col == 6:
        return normalized_address(text)
    if col in ZERO_DECIMAL_DISPLAY_COLUMNS and text:
        try:
            return str(int(Decimal(text).quantize(Decimal("1"), rounding=ROUND_HALF_UP)))
        except InvalidOperation:
            pass
    return " ".join(text.split())


def normalized_address(value: str) -> str:
    text = " ".join(value.lower().split())
    replacements = {
        "drive": "dr",
        "street": "st",
        "road": "rd",
        "court": "ct",
        "avenue": "ave",
        "place": "pl",
        "warehouse": "warehous",
        "loop": "lp",
    }
    for source, target in replacements.items():
        text = re.sub(rf"\b{source}\b", target, text)
    text = re.sub(r"^\s*danze\s+mining\s*,?\s+", "", text)
    text = re.sub(r"\blot\s+\d+\s*\(#(\d+[a-z]?)\)\s*", r"\1 ", text)
    text = re.sub(r"\blot\s+\d+\s*[-,]?\s*", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\b\d{1,2}:\d{2}\s*(?:am|pm)?\b", "", text)
    text = re.sub(r"\bm&b\s+\w+\b", "", text)
    text = re.sub(r"\b(?:am|pm|mb|bt)\b$", "", text)
    text = re.sub(r"[^a-z0-9#]+", " ", text)
    return " ".join(text.split())


def load_rows(path: Path) -> tuple[list[str], list[tuple[int, list[object]]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [display_value(ws.cell(1, col).value) for col in range(1, 25)]
    rows: list[tuple[int, list[object]]] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, 25)]
        if display_value(values[6]):
            rows.append((row_idx, values))
    return headers, rows


def load_source_files(audit_path: Path) -> dict[str, str]:
    source_files: dict[str, str] = {}
    if not audit_path.exists():
        return source_files
    with audit_path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            job = display_value(row.get("job"))
            source_file = display_value(row.get("source_file"))
            if job and source_file:
                source_files[job] = source_file
    return source_files


def load_oversize_reasons(path: Path) -> dict[str, str]:
    reasons: dict[str, str] = {}
    if not path.exists():
        return reasons
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            job = display_value(row.get("job"))
            reason = display_value(row.get("reason"))
            if job and reason:
                reasons[job] = reason
    return reasons


def numeric_value(value: object) -> float:
    text = display_value(value)
    if not text:
        return 0.0
    try:
        return float(Decimal(text))
    except InvalidOperation:
        return 0.0


def generated_has_kd(generated_values: list[object]) -> bool:
    return display_value(generated_values[11]).upper() == "KD" or display_value(generated_values[13]).upper() == "KD"


def is_kd_old_bucket_style(track_values: list[object], generated_values: list[object]) -> bool:
    if not generated_has_kd(generated_values):
        return False
    if numeric_value(track_values[23]):
        return False
    expected_old_v = numeric_value(generated_values[21]) + numeric_value(generated_values[23])
    return abs(numeric_value(track_values[21]) - expected_old_v) < 0.01


def is_goods_pair_order_only_difference(track_values: list[object], generated_values: list[object]) -> bool:
    track_pair = [normalized(track_values[index], index + 1) for index in range(10, 14)]
    generated_pair = [normalized(generated_values[index], index + 1) for index in range(10, 14)]
    if not all(track_pair) or not all(generated_pair):
        return False
    return track_pair == [generated_pair[2], generated_pair[3], generated_pair[0], generated_pair[1]]


def explained_category_for(row: dict[str, str], track_values: list[object], generated_values: list[object]) -> str:
    col = row["col"]
    track = row["track"]
    generated = row["generated"]
    track_norm = row["track_norm"]
    generated_norm = row["generated_norm"]

    if col in {"11", "12", "13", "14"} and is_goods_pair_order_only_difference(track_values, generated_values):
        return "Goods1/Goods2顺序差"
    if col in {"18", "19"} and track_norm and not generated_norm:
        return "手填汇总列：源订单无直接字段"
    if col == "2" and not track_norm and generated_norm:
        return "Job Track漏填PO"
    if col == "3" and (track, generated) in EXPLAINED_BUILDER_PAIRS:
        return "Builder别名"
    if col == "10" and "0.55" in track and "0.6" in generated:
        return "材料旧写法：0.55按源表规则归一为0.6"
    if col in {"21", "22", "24"} and is_kd_old_bucket_style(track_values, generated_values):
        return "KD旧填法：第24列混在第22列"
    return ""


def reason_for(row: dict[str, str], oversize_reasons: dict[str, str]) -> str:
    col = row["col"]
    track = row["track_norm"]
    generated = row["generated_norm"]
    if col == "8":
        return oversize_reasons.get(row["job"], "Over Size与源订单证据不一致，需复核包装关键词或宽度>1024规则")

    if not track and generated:
        base = "Job Track为空，源订单/脚本有值"
    elif track and not generated:
        base = "Job Track有值，脚本未提取到源订单值"
    else:
        base = "两边都有值但不一致"

    if col == "10" and "0.55" in track and "0.6" in generated:
        return f"{base}；0.55mm按源表规则归一为0.6，Job Track仍保留0.55"
    if col == "10" and "0.6" in track and "0.55" in generated:
        return f"{base}；源订单0.55mm未归一到0.6，需复核材料编码规则"
    if col == "20":
        if not track and generated == "0":
            return f"{base}；第20列无MITRE时存在空/0填法差异，常见于KD"
        if track == "0" and not generated:
            return f"{base}；第20列无MITRE时存在空/0填法差异，常见于Door Skin"
        if not track and generated:
            return f"{base}；Job Track未填第20列，脚本按源订单推算MITRE"
        if track and not generated:
            return f"{base}；脚本未能从源订单推算第20列MITRE，可能是特殊版式或手填值"
        return (
            f"{base}；MITRE规则不同：CS按每套14，SPLIT按每套2/双开3，"
            "SPLIT PART A/B ONLY按1，COMMERCIAL普通双开+1，DOUBLE ACTION+0.5，KD不计MITRE"
        )
    if col == "24":
        if not track and generated:
            return f"{base}；源订单为KD/KNOCKDOWN时第24列按KD数量*4计算，Job Track未填"
        if track and not generated:
            return f"{base}；Job Track填了door closer/kd，但源订单未识别到KD数量"
        return f"{base}；door closer/kd按源订单KD数量*4计算"

    column_reasons = {
        "2": "PO号来源或手填值不同",
        "3": "Builder名称/别名规则不同",
        "5": "Zone/区域提取或手填值不同",
        "6": "地址来源不同",
        "10": "Material材料编码或默认材料规则不同",
        "11": "Goods1数量统计不同",
        "12": "Goods1分类不同",
        "13": "Goods2数量统计不同",
        "14": "Goods2分类不同",
        "15": "Ideal D date日期规则不同",
        "16": "Estimate C Date日期规则不同",
        "17": "星期值跟随日期不同",
        "18": "MITRE源值/公式规则不同",
        "19": "Parts源值/公式规则不同",
        "20": "MITRE计算规则不同",
        "21": "Parts公式或五金分桶规则不同",
        "22": "五金数量分桶或倍率规则不同",
        "23": "五金数量分桶或倍率规则不同",
        "24": "door closer/kd规则不同",
    }
    detail = column_reasons.get(col)
    return f"{base}；{detail}" if detail else base


def source_url(source_dir: Path, source_file: str, html_path: Path | None = None) -> str:
    if not source_file:
        return ""
    path = (source_dir / source_file).resolve()
    if html_path is None:
        return path.as_uri()
    relative = os.path.relpath(path, html_path.parent.resolve())
    return quote(Path(relative).as_posix(), safe="/")


def build_diffs(
    generated_path: Path,
    reference_path: Path,
    audit_path: Path,
    oversize_path: Path,
    source_dir: Path,
    html_path: Path | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    headers, generated_rows = load_rows(generated_path)
    _, reference_rows = load_rows(reference_path)
    reference_by_job = {display_value(values[6]): (row_idx, values) for row_idx, values in reference_rows}
    source_files = load_source_files(audit_path)
    oversize_reasons = load_oversize_reasons(oversize_path)

    diffs: list[dict[str, str]] = []
    page_rows: list[dict[str, str]] = []
    for generated_row_idx, generated_values in generated_rows:
        job = display_value(generated_values[6])
        if not job or job not in reference_by_job:
            continue
        track_row_idx, track_values = reference_by_job[job]
        for index in range(24):
            col = index + 1
            track = display_value(track_values[index])
            generated = display_value(generated_values[index])
            track_norm = normalized(track_values[index], col)
            generated_norm = normalized(generated_values[index], col)
            if track_norm == generated_norm:
                continue
            header = headers[index]
            diff = {
                "job": job,
                "generated_row": str(generated_row_idx),
                "track_row": str(track_row_idx),
                "col": str(col),
                "header": header,
                "track": track,
                "generated": generated,
                "track_norm": track_norm,
                "generated_norm": generated_norm,
            }
            diffs.append(diff)

            source_file = source_files.get(job, "")
            page_row = dict(diff)
            page_row.update(
                    {
                        "sourceFile": source_file,
                        "sourceUrl": source_url(source_dir, source_file, html_path),
                        "colKey": f"{col}|{header}",
                        "colLabel": f"{col} - {header}" if header else str(col),
                    }
            )
            category = explained_category_for(diff, track_values, generated_values)
            reason = reason_for(diff, oversize_reasons)
            page_row["explained"] = bool(category)
            page_row["explainCategory"] = category
            page_row["reason"] = f"已解释：{category}；{reason}" if category else reason
            page_rows.append(page_row)
    return diffs, page_rows


def write_diffs(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "job",
        "generated_row",
        "track_row",
        "col",
        "header",
        "track",
        "generated",
        "track_norm",
        "generated_norm",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def render_html(rows: list[dict[str, str]]) -> str:
    payload = json.dumps({"rows": rows}, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>列值差异</title>
<style>
  * {{ box-sizing:border-box; }}
  html {{ min-height:100%; }}
  body {{ min-height:100%; margin:0; overflow-x:hidden; background:#f7f7f5; color:#1f2933; font:14px/1.45 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; }}
  .bar {{ position:sticky; top:0; z-index:5; display:grid; grid-template-columns:minmax(160px,260px) minmax(180px,1fr) minmax(130px,180px) max-content max-content max-content; align-items:center; gap:8px; padding:10px clamp(8px,2vw,16px); background:#fff; border-bottom:1px solid #d8dde3; }}
  input, select, button {{ width:100%; min-width:0; height:36px; border:1px solid #c9d2dc; border-radius:6px; padding:0 10px; background:#fff; font:inherit; }}
  button {{ width:auto; cursor:pointer; white-space:nowrap; }}
  .count {{ align-self:center; justify-self:end; white-space:nowrap; color:#5d6975; text-align:right; }}
  .table-wrap {{ width:100%; max-width:100vw; height:calc(100vh - var(--bar-height,57px)); overflow:auto; -webkit-overflow-scrolling:touch; }}
  table {{ width:100%; min-width:980px; table-layout:fixed; border-collapse:collapse; background:#fff; }}
  th, td {{ border-bottom:1px solid #e3e7ec; padding:8px 10px; text-align:left; vertical-align:top; overflow-wrap:anywhere; word-break:break-word; }}
  th {{ position:sticky; top:0; background:#eef2f5; z-index:4; color:#33414f; font-size:12px; }}
  th:nth-child(1), td:nth-child(1) {{ width:82px; }}
  th:nth-child(2), td:nth-child(2) {{ width:170px; }}
  th:nth-child(3), td:nth-child(3), th:nth-child(4), td:nth-child(4) {{ width:150px; }}
  th:nth-child(5), td:nth-child(5) {{ width:360px; }}
  th:nth-child(6), td:nth-child(6) {{ width:210px; }}
  .job {{ font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; white-space:nowrap; }}
  .col, .file {{ color:#33414f; }}
  .track {{ color:#9d2c2c; font-weight:600; }}
  .script {{ color:#17627b; font-weight:600; }}
  .reason {{ color:#4f5964; }}
  a {{ color:#17627b; text-decoration:none; font-weight:600; }}
  a:hover {{ text-decoration:underline; }}
  @media (max-width: 820px) {{
    body {{ font-size:13px; }}
    .bar {{ grid-template-columns:1fr 1fr; }}
    #q, #col {{ grid-column:1 / -1; }}
    .count {{ justify-self:start; text-align:left; }}
    table {{ min-width:860px; }}
    th, td {{ padding:7px 8px; }}
  }}
  @media (max-width: 520px) {{
    .bar {{ grid-template-columns:1fr; }}
    button, .count {{ width:100%; justify-self:stretch; }}
    .count {{ text-align:left; }}
  }}
</style>
</head>
<body>
  <div class="bar">
    <input id="q" placeholder="搜索Job/值/原因/文件">
    <select id="col"><option value="">全部列</option></select>
    <select id="status"><option value="open">未解释</option><option value="all">全部</option><option value="explained">已解释</option></select>
    <button id="onlyBlank">空值差异</button>
    <button id="clear">清空</button>
    <div class="count" id="count"></div>
  </div>
  <div class="table-wrap" role="region" aria-label="列值差异">
  <table>
    <thead><tr><th>Job</th><th>列</th><th>Job Track</th><th>脚本值</th><th>原因</th><th>原文件</th></tr></thead>
    <tbody id="body"></tbody>
  </table>
  </div>
<script>
const DATA = {payload};
const col = document.getElementById('col');
const status = document.getElementById('status');
let onlyBlank = false;
function esc(v) {{ return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[s])); }}
function val(norm, raw) {{ return norm || raw || ''; }}
function syncBarHeight() {{ document.documentElement.style.setProperty('--bar-height', document.querySelector('.bar').offsetHeight + 'px'); }}
for (const item of [...new Map(DATA.rows.map(r => [r.colKey, r.colLabel])).entries()].sort((a,b) => Number(a[0].split('|')[0])-Number(b[0].split('|')[0]))) {{
  const opt = document.createElement('option');
  opt.value = item[0];
  opt.textContent = item[1];
  col.appendChild(opt);
}}
function rows() {{
  const q = document.getElementById('q').value.trim().toLowerCase();
  return DATA.rows.filter(r => {{
    if (col.value && r.colKey !== col.value) return false;
    if (status.value === 'open' && r.explained) return false;
    if (status.value === 'explained' && !r.explained) return false;
    if (onlyBlank && r.track_norm && r.generated_norm) return false;
    if (!q) return true;
    return [r.job, r.colLabel, r.track_norm, r.generated_norm, r.reason, r.sourceFile].some(v => String(v || '').toLowerCase().includes(q));
  }});
}}
function render() {{
  syncBarHeight();
  const rs = rows();
  document.getElementById('count').textContent = rs.length + ' 条';
  document.getElementById('onlyBlank').style.background = onlyBlank ? '#e5f0ea' : '#fff';
  document.getElementById('body').innerHTML = rs.map(r => `<tr><td class="job">${{esc(r.job)}}</td><td class="col">${{esc(r.colLabel)}}</td><td class="track">${{esc(val(r.track_norm, r.track))}}</td><td class="script">${{esc(val(r.generated_norm, r.generated))}}</td><td class="reason">${{esc(r.reason)}}</td><td class="file">${{r.sourceUrl ? `<a href="${{esc(r.sourceUrl)}}">${{esc(r.sourceFile)}}</a>` : esc(r.sourceFile)}}</td></tr>`).join('');
}}
document.getElementById('q').addEventListener('input', render);
col.addEventListener('change', render);
status.addEventListener('change', render);
document.getElementById('onlyBlank').addEventListener('click', () => {{ onlyBlank = !onlyBlank; render(); }});
document.getElementById('clear').addEventListener('click', () => {{ document.getElementById('q').value=''; col.value=''; status.value='open'; onlyBlank=false; render(); }});
window.addEventListener('resize', syncBarHeight);
render();
</script>
</body>
</html>
"""


def write_html(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(render_html(rows), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Regenerate Job Track comparison CSV and filterable HTML report.")
    parser.add_argument("--generated", type=Path, default=DEFAULT_GENERATED)
    parser.add_argument("--reference", type=Path, default=DEFAULT_REFERENCE)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--oversize", type=Path, default=DEFAULT_OVERSIZE)
    parser.add_argument("--diffs", type=Path, default=DEFAULT_DIFFS)
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML)
    parser.add_argument("--report-html", type=Path, default=DEFAULT_REPORT_HTML)
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    args = parser.parse_args()

    diffs, page_rows = build_diffs(args.generated, args.reference, args.audit, args.oversize, args.source_dir, args.html)
    write_diffs(args.diffs, diffs)
    write_html(args.html, page_rows)
    write_html(args.report_html, page_rows)

    material_diffs = sum(1 for row in diffs if row["col"] == "10")
    oversize_diffs = sum(1 for row in diffs if row["col"] == "8")
    print(f"Wrote {len(diffs)} diffs to {args.diffs}")
    print(f"Material diffs: {material_diffs}")
    print(f"Over Size diffs: {oversize_diffs}")
    print(f"Wrote HTML to {args.html}")
    print(f"Wrote report HTML to {args.report_html}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
