from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook

import extract


def save_workbook(wb: Workbook, path: Path) -> Path:
    wb.save(path)
    return path


def test_material_code_preserves_0_55_galv() -> None:
    assert extract.material_code("0.55mm Galv") == "0.55G"


def test_material_code_preserves_0_55_colorbond_colours() -> None:
    assert extract.material_code("0.55mm Monument") == "0.55CB"
    assert extract.material_code("0.55mm Surfmist") == "0.55CB"
    assert extract.material_code("0.55mm Dover White") == "0.55CB"


def test_material_code_reads_short_zincanneal_and_s_steel() -> None:
    assert extract.material_code("1.2 ZA") == "1.2Z"
    assert extract.material_code("0.80mm S/Steel Rimex Sheet") == "0.8SS"


def test_builder_aliases_match_job_track_names() -> None:
    cases = {
        "Dale Alcock Projects": "DALE ALCOCK",
        "BW RESIDENTIAL": "BW",
        "M&B": "M & B",
        "M&B Sales": "M&B",
        "Pyramid Construction": "PYRAMID",
        "Fortitude Living P/L & GLiving P/L": "FORTITUDE",
        "MURRAY RIVER NORTH PTY LTD": "MRN",
        "Carnarvon Timber & Hardware": "CARNARVON",
        "Atlas Building (WA) Pty Ltd": "ATLAS",
        "Access Projects & Construction Pty Ltd": "ACCESS",
        "Built": "BUILT",
        "Project Building Supplies": "PROJECT",
        "Velocity Building": "VELOCITY",
        "McCorkell Construction": "McCorkell",
        "EvaBuilt Construction": "EvaBuilt",
        "Johns Building Supplies Pty Ltd": "JOHNS BUILDING",
    }
    for raw, expected in cases.items():
        row = extract.ExtractedRow()
        assert extract.normalize_builder(raw, row) == expected
        assert row.manual_check == []


def test_deluxe_split_profile_counts_as_split_not_split_dl(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    ws["A11"] = "1.05mm Zincanneal"
    ws["C11"] = 2
    ws["D11"] = "Deluxe Split"
    ws["F11"] = 2060
    ws["G11"] = 823
    ws["H11"] = "RIGHT"
    ws["I11"] = 2
    ws["J11"] = "WELDED"
    ws["K11"] = "S1"
    ws["L11"] = 1000

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "deluxe-split.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "SPLIT"]


def test_previous_business_day_uses_wa_weekday_not_china_may_holiday() -> None:
    assert extract.previous_business_day(dt.date(2026, 5, 5)) == dt.date(2026, 5, 4)


def test_previous_business_day_skips_wa_day_public_holiday() -> None:
    assert extract.previous_business_day(dt.date(2026, 6, 2)) == dt.date(2026, 5, 29)


def test_parse_date_handles_time_three_digit_year_and_zero_month() -> None:
    assert extract.parse_date("16/04/2026 10.00AM") == dt.date(2026, 4, 16)
    assert extract.parse_date("27/05/206") == dt.date(2026, 5, 27)
    assert extract.parse_date("23/0/2026") == dt.date(2026, 1, 23)


def test_door_skin_main_sheet_splits_skin_and_commercial_profile_rows(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27063 REV00"
    ws["A2"] = "Builder:"
    ws["B2"] = "Danze Mining"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 1, 16)
    ws["A6"] = "Purchase Order:"
    ws["B6"] = "02-6228-01"
    ws["A7"] = "Zone"
    ws["B7"] = "57d"

    for col, value in enumerate(["MATERIAL", "PROFILE", "QUANTITY", "WIDTH", "LENGTH", "BOLT"], start=1):
        ws.cell(11, col).value = value
    skin_rows = [
        ("1.05mm Zinc", "SKIN - 1.05MM -BLANK (44)", 1, 871, 2079),
        ("1.05mm Zinc", "SKIN - 1.05MM BLANK (16)", 1, 868, 2076),
        ("1.05mm Zinc", "SKIN - 1.05MM -BLANK (44)", 1, 1027, 2244),
        ("1.05mm Zinc", "SKIN - 1.05MM BLANK (16)", 1, 1024, 2241),
    ]
    for row_idx, values in enumerate(skin_rows, start=12):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    ws["A18"] = "Material:"
    ws["B18"] = "1.2mm"
    ws["C18"] = "Zinc"
    ws["A20"] = "Quantity"
    ws["C20"] = "OVERALL"
    ws["D20"] = "OVERALL"
    ws["E20"] = "FOUR SIDED"
    ws["B21"] = "PROFILE"
    ws["C21"] = "HEIGHT"
    ws["D21"] = "WIDTH"
    ws["E21"] = "FRAME"
    for row_idx, values in {
        22: [2, "A", 500, 500, "YES"],
        23: [4, "A", 730, 630, "YES"],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(
        save_workbook(wb, tmp_path / "Danze Mining 27063 Door Skins 1.05 Blank.xlsx"),
        infer_manual=True,
    )

    assert row.values[10:14] == [6, "COMMERCIAL", 4, "DS"]


def test_profileless_main_sheet_keeps_concealed_rows_separate(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 28065 REV00"
    ws["A2"] = "Builder:"
    ws["B2"] = "Orthonova"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 3, 2)
    ws["A6"] = "Purchase Order:"
    ws["B6"] = "PO-28065"
    ws["A9"] = "Material:"
    ws["B9"] = "1.05mm Zincanneal ZF100"
    for col, value in enumerate(
        [None, "WALL", "FRAME TYPE", "DOOR", "REVEAL", "REVEAL", None, "HINGE"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "TYPE", "(PER A_6020)", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY"],
        start=1,
    ):
        ws.cell(12, col).value = value
    for row_idx, values in {
        13: ["D01", "P-01", "16 (Concealed)", "40mm", 2360, 1646, "DOUBLE", 1],
        14: ["D02", "P-02", "16", "40mm", 2360, 1346, "DOUBLE", 1],
        15: ["D03", "P-02", "16", "40mm", 2360, 1346, "DOUBLE", 1],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "concealed-main-sheet.xlsx"), infer_manual=True)

    assert row.values[10:14] == [2, "COMMERCIAL", 1, "CONCEALED"]


def test_standard_worksheet_splits_concealed_cavity_slider_rows(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    for row_idx, values in {
        11: ["1.05mm Zincanneal", "", 2, "CONCEALED FRAME", "85-120", 2060, 823, "LEFT", 2, "SCREW FIXED PREP", "S1", 1000, "NO", "NO", "NO"],
        12: ["1.05mm Zincanneal", "", 1, "CONCEALED CAVITY SLIDER", "114", 2360, 700, 750, 1490, "BRIO", "N", "N", "STUD", "N/A", ""],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "concealed-cavity.xlsx"), infer_manual=True)

    assert row.values[10:14] == [2, "CONCEALED", 1, "CS"]


def test_standard_worksheet_keeps_bauhaus_as_goods_type(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    for row_idx, values in {
        11: ["1.05mm Zincanneal", "", 12, "Bauhaus", "95", 2360, 823, "LEFT"],
        12: ["1.05mm Zincanneal", "", 1, "Modern", "95", 2360, 1640, ""],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "bauhaus.xlsx"), infer_manual=True)

    assert row.values[10:14] == [12, "Bauhaus", 1, "MODERN"]


def test_sheet1_dyna_profile_counts_as_commercial(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 27838"
    ws["A2"] = "Builder:"
    ws["B2"] = "Buildform"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 2, 2)
    ws["A12"] = "Material:"
    ws["B12"] = "1.2mm Zinc"
    for col, value in enumerate(["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "HAND", "HINGE QTY"], start=1):
        ws.cell(14, col).value = value
    for col, value in enumerate([1, "SPLIT A + B (40mm Door)", 2060, 823, "RIGHT", 3], start=1):
        ws.cell(15, col).value = value
    ws["A31"] = "Material:"
    ws["B31"] = "1.2mm Galv"
    for col, value in enumerate(["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "HAND", "HINGE QTY"], start=1):
        ws.cell(33, col).value = value
    for col, value in enumerate([1, "DYNA (40mm Door)", 2090, 975, "LEFT", 3], start=1):
        ws.cell(34, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "buildform-dyna.xlsx"), infer_manual=True)

    assert row.values[10:14] == [1, "SPLIT", 1, "COMMERCIAL"]


def test_short_code_profiles_count_as_commercial(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 26944"
    ws["A12"] = "Material:"
    ws["B12"] = "1.2mm Zinc"
    for col, value in enumerate(["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "HAND", "HINGE QTY"], start=1):
        ws.cell(14, col).value = value
    for row_idx, profile in enumerate(["SD01", "IW-FR1", "125 CAST (40mm Door)", "A + B", "HYDE-1603-FR-001"], start=15):
        values = [f"D{row_idx}", profile, 2060, 923, "LEFT", 3]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "short-code-commercial.xlsx"), infer_manual=True)

    assert row.values[10:12] == [5, "COMMERCIAL"]


def test_main_sheet_ignores_offset_brick_ties_after_material_section(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27376"
    ws["A2"] = "Builder:"
    ws["B2"] = "Australian Fire Door Company"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 1, 30)
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY"], start=1):
        ws.cell(12, col).value = value
    for col, value in enumerate(["D01-FR SCULLERY", "A", 2100, 1020, "LEFT", 4], start=1):
        ws.cell(13, col).value = value
    ws["A16"] = "Material:"
    ws["B16"] = "Stainless Steel"
    ws["A18"] = "OFFSET BRICK TIES"
    ws["B18"] = 10

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "AFDC 27376 Brick.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_main_sheet_frame_type_rows_count_commercial_without_product_qty(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27033"
    for col, value in enumerate(["Door #", "Frame", "Wall", "REVEAL", "REVEAL", "", "DOOR", "HINGE"], start=1):
        ws.cell(11, col).value = value
    for col, value in enumerate(["", "Type", "Type", "HEIGHT", "WIDTH", "HAND", "TYPE", "QTY"], start=1):
        ws.cell(12, col).value = value
    rows = {
        13: ["D1/1.03", "SD04", "SP-01 + LINING", 2360, 925, "RIGHT", "SC = 40mm", 4],
        14: ["D1/1.04", "SD04", "SP-01 + LINING", 2360, 1386, "UNEVEN DOUBLE", "SC = 40mm", "8 (4 EACH SIDE)"],
        15: ["D9 1.01", "SD07", "SP-01", 2360, 725, "RIGHT", "SC = 40mm", 4],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "smc-frame-type.xlsx"), infer_manual=True)

    assert row.values[10:12] == [3, "COMMERCIAL"]


def test_main_sheet_uses_later_parseable_delivery_date(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27075"
    ws["A4"] = "Delivery Date:"
    ws["B4"] = "Dunreath Drive, Perth Airport"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.datetime(2026, 1, 16)
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY"], start=1):
        ws.cell(12, col).value = value
    for col, value in enumerate(["D1", "A", 2100, 900, "RIGHT", 1], start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-duplicate-date.xlsx"), infer_manual=True)

    assert row.values[14:17] == [dt.date(2026, 1, 16), dt.date(2026, 1, 15), "Thursday"]


def test_main_sheet_frame_type_double_rows_add_smc_hardware_parts(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 27033 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        ["Door #", "Frame", "Wall", "REVEAL", "REVEAL", None, "DOOR", "HINGE", "BACKING PLATE", "STRIKER", "STRIKER", "DYNA"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        [None, "Type", "Type", "HEIGHT", "WIDTH", "HAND", "TYPE", "QTY", "TO SUIT", "TYPE", "HEIGHT", "BOLT"],
        start=1,
    ):
        ws.cell(12, col).value = value
    rows = {
        13: ["D1/1.03", "SD04", "SP-01 + LINING", 2360, 925, "RIGHT", "SC = 40mm", 4, "100X75X2.5", "MORTICE LOCK (S1)", 1032, None],
        14: ["D1/1.04", "SD04", "SP-01 + LINING", 2360, 1386, "UNEVEN DOUBLE", "SC = 40mm", "8 (4 EACH SIDE)", "100X75X2.5", "-", "-", None],
        15: ["D9 1.01", "SD07", "SP-01", 2360, 1646, "DOUBLE", "SC = 40mm", "6 (3 each side)", "100x75x2.5", "-", "-", None],
        16: ["D10 1.01", "SD07", "SP-01", 2360, 503, "LEFT", "SC = 40mm", 3, "BELLEVUE CEAM BAC1230SSI", "MORTICE LOCK (S1)", 1032, "-"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "smc-frame-type-double.xlsx"), infer_manual=True)

    assert row.values[10:12] == [4, "COMMERCIAL"]
    assert row.values[19:24] == [5, 34, 4, 21, None]


def test_main_sheet_frame_type_overall_height_adds_smc_sill_part(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 27197 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        ["Door #", "Frame", "Wall", "OVERALL", "REVEAL", None, "DOOR", "HINGE", "BACKING PLATE", "STRIKER", "STRIKER", "DYNA"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        [None, "Type", "Type", "HEIGHT", "WIDTH", "HAND", "TYPE", "QTY", "TO SUIT", "TYPE", "HEIGHT", "BOLT"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["D1/1.19", "SD01", "SP-01", 2150, 923, "LEFT", "SC = 40mm", 4, "100X75X2.5", "-", "-", "-"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "smc-overall-height.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 7, 1, 4, None]


def test_sheet1_compound_a_b_profile_counts_one_commercial_row(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 30150"
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE"],
        start=1,
    ):
        ws.cell(12, col).value = value
    for col, value in enumerate(
        ["L1-077 D1002", "A + B", 2360, 974, "RIGHT", 4, "100X100X2.5", "ES2100"],
        start=1,
    ):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Spence 30150 Stud.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_sheet1_compound_c_d_profile_counts_as_commercial(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 30150"
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE"],
        start=1,
    ):
        ws.cell(12, col).value = value
    for col, value in enumerate(
        ["GF-079-D0002", "C + D", 2360, 974, "RIGHT", 4, "100X100X2.5", "ES2100"],
        start=1,
    ):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Spence 30150 Stud.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_part_b_only_profile_counts_as_split(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 27963"
    ws["A12"] = "Material:"
    ws["B12"] = "1.05mm Zinc"
    for col, value in enumerate(["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "HAND", "HINGE QTY"], start=1):
        ws.cell(14, col).value = value
    for row_idx in range(15, 20):
        values = [f"D{row_idx}", "PART B ONLY (165 WALL)", 2060, 923, "RIGHT", 3]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "part-b-only.xlsx"), infer_manual=True)

    assert row.values[10:12] == [5, "SPLIT"]


def test_nonstandard_worksheet_profile_qty_counts_mb(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["A1"] = "AUSMET JOB # 27338"
    for col, value in enumerate(["Material", "Stock", "Qty", "Profile"], start=1):
        ws.cell(9, col).value = value
    for col, value in enumerate(
        ["Other", None, 15, "Combination Gas/Electric (Non-Rebated)"],
        start=1,
    ):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-mb.xlsx"), infer_manual=True)

    assert row.values[10:12] == [15, "MB"]


def test_standard_worksheet_other_mb_does_not_become_second_goods(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["C1"] = 29053
    for col, value in enumerate(["Material", "Stock", "Qty", "Profile"], start=1):
        ws.cell(9, col).value = value
    for col, value in enumerate(["1.2mm Zinc", None, 1, "A"], start=1):
        ws.cell(10, col).value = value
    for col, value in enumerate(
        ["Other", None, 15, "Combination Gas/Electric (Non-Rebated)"],
        start=1,
    ):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-mb-accessory.xlsx"), infer_manual=True)

    assert row.values[10:14] == [1, "COMMERCIAL", None, None]


def test_flat_strips_note_counts_as_fs(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27992 REV00"
    ws["A10"] = "FLAT STRIPS (1.2 OR 1.6) 50MM WIDE X 2150MM L   QTY = 8 LENGTHS."

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "flat-strips.xlsx"), infer_manual=True)

    assert row.values[10:12] == [8, "FS"]


def test_trad_dyna_material_table_splits_door_skin_quantity(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 27075"
    ws["A2"] = "Builder:"
    ws["B2"] = "AFDC"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 1, 20)
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm"
    ws["C9"] = "Zinc"
    for col, value in enumerate(["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY"], start=1):
        ws.cell(12, col).value = value
    for row_idx, values in {
        13: ["ST2.B", "A", 2080, 990, "RIGHT", 4],
        14: ["ST2.C", "A", 2080, 990, "RIGHT", 4],
        15: ["G04.A", "D", 2400, 1700, "DOUBLE", "8 (4 EACH SIDE)"],
        16: ["G08.A", "D", 2100, 1695, "DOUBLE", "8 (4 EACH SIDE)"],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    for col, value in enumerate(["MATERIAL", "PROFILE", "QUANTITY", "LENGTH", "C/SINK HOLES"], start=1):
        ws.cell(20, col).value = value
    for row_idx, values in {
        21: ["1.6mm Zinc", "B", 2, 2100, "6 - EVENLY SPACED"],
        22: ["1.6mm Zinc", "B", 1, 2000, "6 - EVENLY SPACED"],
    }.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(
        save_workbook(wb, tmp_path / "AFDC 27075 Trad Dyna Dbl Rebate.xlsx"),
        infer_manual=True,
    )

    assert row.values[10:14] == [4, "COMMERCIAL", 3, "DS"]


def test_total_doors_nonstandard_worksheet_splits_door_and_flat_sheet(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    wb.create_sheet("Data")
    ws["A1"] = "Ausmet Job #"
    ws["C1"] = 29306
    ws["A2"] = "Builder:"
    ws["C2"] = "Cash Sale"
    ws["A4"] = "Delivery Address:"
    ws["C4"] = "PICK UP"
    ws["A5"] = "Delivery Date:"
    ws["C5"] = dt.date(2026, 5, 14)
    ws["A6"] = "PO No:"
    ws["C6"] = "TOTAL DOORS"
    for col, value in enumerate(
        [
            "Material",
            "Door #",
            "Qty",
            "WALL TYPE",
            "FRAME TYPE",
            "DOOR THICKNESS",
            "OVERALL HEIGHT",
            "OVERALL WIDTH",
            "HAND",
            "HINGE QTY",
            "BACKING PLATE TO SUIT",
            "STRIKER TYPE",
            "STRIKER HEIGHT",
            "CSK HOLES A",
            "Format Tag",
        ],
        start=1,
    ):
        ws.cell(10, col).value = value
    ws["A11"] = "1.6mm zinc"
    for col, value in {
        3: 1,
        6: "40mm",
        7: 2115,
        8: 943,
        9: "LEFT",
        10: 3,
        11: "100X75X2.5",
        12: "S1",
        13: 1000,
    }.items():
        ws.cell(12, col).value = value
    ws["A16"] = "GALV"
    ws["C16"] = "Qty"
    ws["F16"] = "WIDTH"
    ws["G16"] = "LENGTH"
    ws["C17"] = 2
    ws["F17"] = 840
    ws["G17"] = 2030
    ws["H17"] = "FLAT SHEET"
    ws["C18"] = 2
    ws["G18"] = 2400
    ws["H18"] = "CAPPING"

    row = extract.extract_workbook(
        save_workbook(wb, tmp_path / "29306 - CASH SALE - TOTAL DOORS.xlsx"),
        infer_manual=True,
    )

    assert row.values[10:14] == [1, "SPLIT", 2, "DS"]


def extracted_row(job: str, source_file: str, marker: str) -> extract.ExtractedRow:
    row = extract.ExtractedRow(source_file=source_file)
    row.values[6] = job
    row.values[19] = marker
    return row


def test_row_with_only_job_number_is_not_order_payload() -> None:
    row = extract.ExtractedRow(source_file="CR PL600325装箱单.xlsx")
    row.values[6] = 14149

    assert not extract.row_has_order_payload(row)


def test_row_with_order_fields_is_order_payload() -> None:
    row = extract.ExtractedRow(source_file="29306 - CASH SALE - TOTAL DOORS.xlsx")
    row.values[6] = 29306
    row.values[2] = "CASH SALE"

    assert extract.row_has_order_payload(row)


def test_duplicate_jobs_keep_highest_source_version(tmp_path: Path) -> None:
    older_path = tmp_path / "29698__0178__old__29698 BEYOND RES SPLIT + CS.xlsx"
    newer_path = tmp_path / "29698__0216__new__29698 BEYOND RES SPLIT + CS.xlsx"
    older_path.touch()
    newer_path.touch()
    older = extracted_row("29698", older_path.name, "old")
    newer = extracted_row("29698", newer_path.name, "new")

    rows = extract.dedupe_latest_rows([older, newer], [older_path, newer_path])

    assert rows == [newer]


def test_duplicate_jobs_without_source_version_use_latest_file_mtime(tmp_path: Path) -> None:
    older_path = tmp_path / "29782 APEX MKD + ST.xlsx"
    newer_path = tmp_path / "29782 APEX SPLIT.xlsx"
    older_path.touch()
    newer_path.touch()
    older_time = 1_700_000_000
    newer_time = 1_700_000_100
    older_path.touch()
    newer_path.touch()
    import os

    os.utime(older_path, (older_time, older_time))
    os.utime(newer_path, (newer_time, newer_time))
    older = extracted_row("29782", older_path.name, "old")
    newer = extracted_row("29782", newer_path.name, "new")

    rows = extract.dedupe_latest_rows([older, newer], [older_path, newer_path])

    assert rows == [newer]


def test_duplicate_jobs_without_source_version_use_email_uid_before_mtime(tmp_path: Path) -> None:
    older_path = tmp_path / "INBOX-016837-00016837-3-29698 BEYOND RES SPLIT + CS.xlsx"
    newer_path = tmp_path / "INBOX-016904-00016904-2-29698 BEYOND RES SPLIT + CS.xlsx"
    older_path.touch()
    newer_path.touch()

    import os

    os.utime(older_path, (1_700_000_100, 1_700_000_100))
    os.utime(newer_path, (1_700_000_000, 1_700_000_000))
    older = extracted_row("29698", older_path.name, "old")
    newer = extracted_row("29698", newer_path.name, "new")

    rows = extract.dedupe_latest_rows([older, newer], [older_path, newer_path])

    assert rows == [newer]


def worksheet_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    wb.create_sheet("Data")
    ws["C1"] = 99999
    ws["C2"] = "Test Builder"
    ws["C5"] = "2026-06-15"
    headers = [
        "Material",
        "Stock",
        "Qty",
        "Profile",
        "B/O",
        "Reveal Height",
        "Reveal Width",
        "Hand",
        "Hinge Qty",
        "Hinge Type",
        "Striker Type",
        "Striker Height",
        "Sill",
        "Slider",
        "Double",
        "CL1",
        "CL3",
        "Striker Type2",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(9, col).value = value
    return wb


def test_standard_worksheet_hinge_qty_is_not_over_size_width(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = ["1.6mm Zinc", "", 1, "Commercial", "", 2060, 923, "RIGHT", 1670, "HINGE PREP", "S1", 1000]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-commercial-hinge-qty.xlsx"), infer_manual=True)

    assert row.values[7] is None


def test_standard_worksheet_reveal_width_uses_1220_over_size_threshold(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    rows = {
        11: ["1.05mm Zincanneal", "", 1, "Modern", "114", 2060, 1203, "RIGHT", 4, "WELDED", "-", "-"],
        12: ["1.05mm Zincanneal", "", 2, "Modern", "114", 2060, 1220, "LEFT", 4, "WELDED", "-", "-"],
        13: ["1.05mm Zincanneal", "", 3, "Modern", "114", 2060, 1221, "RIGHT", 4, "WELDED", "-", "-"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-reveal-threshold.xlsx"), infer_manual=True)

    assert row.values[7] == 3


def test_standard_worksheet_scans_material_from_zero_qty_rows_and_notes(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    ws["I4"] = "ATTACH STRIKER PLATES FRAMES. 1.2 ZA"
    rows = {
        10: ["1.2mm\nZincanneal", "", 0, "", "", 0, 0, "", "", "", "", ""],
        11: ["", "", 5, "Modern Knockdown - Screw Fix", "118", 2060, 923, "RIGHT", 3, "SCREW FIX PREP", "S1", 1032],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-zero-material.xlsx"), infer_manual=True)

    assert row.values[9] == "1.2Z"


def test_standard_worksheet_screw_fix_kd_uses_seven_x_parts_per_qty(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    rows = {
        11: [
            "1.05mm Zincanneal",
            "",
            1,
            "Modern Knock Down - Screw Fix",
            "114",
            2060,
            723,
            "LEFT",
            2,
            "WELDED",
            "S1",
            1000,
            "NO",
            "NO",
            "NO",
        ],
        12: [
            "1.05mm Zincanneal",
            "",
            2,
            "Modern Knock Down - Screw Fix",
            "114",
            2060,
            723,
            "LEFT WC",
            2,
            "WELDED",
            "S1",
            1000,
            "NO",
            "NO",
            "NO",
        ],
        13: [
            "1.05mm Zincanneal",
            "",
            3,
            "Modern Knock Down - Screw Fix",
            "114",
            2060,
            823,
            "RIGHT",
            2,
            "WELDED",
            "S1",
            1000,
            "NO",
            "NO",
            "NO",
        ],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-kd-screw-fix.xlsx"), infer_manual=True)

    assert row.values[11] == "KD"
    assert row.values[23] == 42


def test_standard_worksheet_plain_hinge_stays_in_v_bucket(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Split",
        "85-125",
        2060,
        923,
        "RIGHT",
        "3",
        "HINGE PREP",
        "S1+RDL",
        "1000",
        "NO",
        "NO",
        "NO",
        "Split",
        "85-125",
        "S1+S1",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "plain.xlsx"), infer_manual=True)

    assert row.values[19:24] == [2, 5, 5, None, None]


def test_standard_worksheet_screw_fixed_prep_uses_w_bucket_and_skips_other_rows(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    rows = {
        11: [
            "1.05mm Zincanneal",
            "",
            1,
            "Modern",
            "95",
            2360,
            823,
            "LEFT WC",
            "3",
            "SCREW FIXED PREP",
            "S1+ZANDA 10421",
            "1000 + 1150",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            1,
        ],
        12: [
            "1.05mm Zincanneal",
            "",
            1,
            "Modern",
            "95",
            2360,
            823,
            "LEFT",
            "3",
            "SCREW FIXED PREP",
            "S1",
            "1000",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            1,
        ],
        14: [
            "Other",
            "",
            1,
            "Single Elec with Lock & View (Non-Rebated)",
            None,
            0,
            0,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            None,
            None,
            1,
        ],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "screw-fixed.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "MODERN"]
    assert row.values[19:24] == [2, 9, 3, 6, None]


def test_standard_worksheet_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Modern",
        "95",
        2060,
        823,
        "RIGHT",
        "2",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Modern",
        "95",
        "S1",
    ]
    for row_idx in (11, 12):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[12].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-worksheet.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "MODERN"]
    assert row.values[19:24] == [1, 3, 3, None, None]


def test_standard_worksheet_negative_qty_does_not_cancel_goods_qty(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    for row_idx, qty in ((11, -1), (12, 1)):
        values = [
            "1.05mm Zincanneal",
            "",
            qty,
            "Modern",
            "95",
            2060,
            823,
            "RIGHT",
            "2",
            "WELDED",
            "S1",
            "1000",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            "S1",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "negative-qty.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "MODERN"]
    assert row.values[19] == 1


def test_standard_worksheet_deluxe_cleats_ignores_negative_return_parts(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    for row_idx, qty in ((11, -1), (12, 1)):
        values = [
            "1.05mm Zincanneal",
            "",
            qty,
            "Deluxe Dry Lining",
            "114",
            2060,
            723,
            "RIGHT",
            "2",
            "WELDED",
            "S1",
            "1000",
            "NO",
            "NO",
            "NO",
            "Deluxe Dry Lining",
            "114",
            "S1",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "30446 PRIME DELUXE DRY LINING + CLEATS.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "DELUXE"]
    assert row.values[21] == 9


def test_standard_worksheet_kd_parts_go_to_x_bucket_not_v_bucket(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        2,
        "Modern Knock Down",
        "114",
        2060,
        823,
        "RIGHT",
        "3",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Modern Knock Down",
        "114",
        "S1",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-modern-kd.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "KD"]
    assert row.values[19:24] == [0, 11, 8, None, 8]


def test_standard_worksheet_equal_qty_keeps_source_goods_order(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    split_values = [
        "1.05mm Zincanneal",
        "",
        8,
        "Split",
        "85-125",
        2060,
        823,
        "RIGHT",
        "3",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Split",
        "85-125",
        "S1",
    ]
    for col, value in enumerate(split_values, start=1):
        ws.cell(11, col).value = value
    ws.cell(13, 1).value = "Cavity Sliders"
    cavity_values = [
        "",
        "",
        8,
        "Modern",
        "114",
        2060,
        700,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    for col, value in enumerate(cavity_values, start=1):
        ws.cell(14, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "equal-goods-order.xlsx"), infer_manual=True)

    assert row.values[10:14] == [8, "SPLIT", 8, "CS"]
    assert row.values[19] == 128


def test_standard_worksheet_cavity_soft_closer_does_not_count_as_cs_goods(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    ws.cell(12, 1).value = "Cavity Sliders"
    rows = {
        13: ["", "", 1, "Modern", "114", 2360, 700, "", "", "", "", "", "", "", "", "", "", ""],
        14: ["", "", 1, "Brio Soft Closer (bev to deliver)", "", 0, "", "", "", "", "", "", "", "", "", "", "", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "cavity-soft-closer.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "CS"]
    assert row.values[19] == 14


def sheet1_profile_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Job 30095"
    headers = ["QTY", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGES", "TYPE", "HEIGHT", "HOLES"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    return wb


def test_sheet1_mortar_guards_go_to_w_without_hinge_qty(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Ausmet Job # 27094"
    ws["A11"] = "Material"
    ws["B11"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "OVERALL", "OVERALL", None, "HINGE", None, "STRIKER", "STRIKER", "MORTAR", "DOOR", "BRICK"],
        start=1,
    ):
        ws.cell(13, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGE TYPE", "TYPE", "HEIGHT", "GUARDS", "CLOSER", "TIES"],
        start=1,
    ):
        ws.cell(14, col).value = value
    rows = {
        15: ["D.03/T1.42", 38, 2057, 1050, "LEFT", 4, "100X75X2.5", "S1", 1032, "YES", "DC1", "1 Pack Wire Ties"],
        16: ["D.05/T1.43", 36, 2057, 1250, "DOUBLE", "8 (4 EACH SIDE)", "100X75X2.5", "-", "-", "YES", "-", "1 Pack Wire Ties"],
        17: ["D.03/T1.51", 38, 2057, 1050, "RIGHT", 4, "100X75X2.5", "S1 + MIB", "S1 @ 950 + MIB @ 1050", "YES", "DC1", "1 Pack Wire Ties"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-mortar-guards.xlsx"), infer_manual=True)

    assert row.values[10:12] == [3, "COMMERCIAL"]
    assert row.values[21:23] == [5, 12]
    assert row.values[20] == 22


def thomas_sheet1_knockdown_cavity_book(delivery_date: dt.datetime) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Ausmet Job # 27035"
    ws["A3"] = "Builder:"
    ws["B3"] = "Thomas Building"
    ws["A5"] = "Delivery Address:"
    ws["B5"] = "Pick Up, by Bishops Transport"
    ws["A6"] = "Delivery Date:"
    ws["B6"] = delivery_date
    ws["A7"] = "PO:"
    ws["B7"] = "946-100101"
    ws["A11"] = "Material"
    ws["B11"] = "1.2mm Zinc"
    for col, value in enumerate(
        ["DOOR", None, "REVEAL", "REVEAL ", None, "HINGE ", "WELDED", "STRIKER ", "STRIKER "],
        start=1,
    ):
        ws.cell(12, col).value = value
    for col, value in enumerate(
        ["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGES", "TYPE", "HEIGHT"],
        start=1,
    ):
        ws.cell(13, col).value = value
    hinge_qtys = [3, "4 (2 EACH SIDE)", 2, 2, 2, 2, 2]
    for offset, hinge_qty in enumerate(hinge_qtys):
        row_idx = 14 + offset
        ws.cell(row_idx, 1).value = f"L511 COOLERIN - {offset + 1}"
        ws.cell(row_idx, 2).value = "125mm B/O MODERN KNOCK DOWN-SCREW FIXED MITRE"
        ws.cell(row_idx, 3).value = 2060
        ws.cell(row_idx, 4).value = 823
        ws.cell(row_idx, 5).value = "DOUBLE" if offset == 1 else "RIGHT"
        ws.cell(row_idx, 6).value = hinge_qty
        ws.cell(row_idx, 7).value = "WELDED - 100X75X1.6"
        ws.cell(row_idx, 8).value = "S1"
        ws.cell(row_idx, 9).value = 1000

    ws["A23"] = "Material"
    ws["B23"] = "1.6mm Zinc"
    for col, value in enumerate(
        [
            "DOOR",
            None,
            "REVEAL",
            "REVEAL ",
            "POCKET",
            "OVERALL",
            "STRIKER",
            "STRIKER",
            "HARDWARE",
            "INFILL",
            "PELMET",
            "FLUSH STUD",
        ],
        start=1,
    ):
        ws.cell(24, col).value = value
    for col, value in enumerate(
        ["NUMBER", "PROFILE", "HEIGHT", "WIDTH", "WIDTH", "WIDTH", "TYPE", "HEIGHT", None, None, None, "BRACKETS"],
        start=1,
    ):
        ws.cell(25, col).value = value
    for offset, room in enumerate(["ENS", "WIR", "BATH"]):
        row_idx = 26 + offset
        values = [
            f"L511 COOLERIN - {room}",
            "112mm B/O Modern Cavity Slider",
            2060,
            800,
            850,
            1690,
            "Lockwood 7400SPDP",
            1000,
            "COWDROY",
            "YES",
            "YES",
            "7 (3 TO CLOSING JAMB + 4 TO HEAD)",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    return wb


def test_sheet1_reveal_width_1203_is_not_over_size(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Job 30095"
    ws["A6"] = "Material"
    ws["B6"] = "1.05mm Zincanneal"
    headers = [
        "Door #",
        "PROFILE",
        "OVERALL HEIGHT",
        "REVEAL WIDTH",
        "HAND",
        "HINGE QTY",
        "WELDED HINGES",
        "STRIKER TYPE",
        "STRIKER HEIGHT",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    rows = {
        13: ["D1", "A", 2229, 1203, "RIGHT", 4, "100X75X1.6", "-", "-"],
        14: ["D2", "A", 2229, 1220, "LEFT", 4, "100X75X1.6", "-", "-"],
        15: ["D3", "A", 2229, 1221, "RIGHT", 4, "100X75X1.6", "-", "-"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-reveal-threshold.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_sheet1_thomas_early_knockdown_cavity_batch_matches_track(tmp_path: Path) -> None:
    wb = thomas_sheet1_knockdown_cavity_book(dt.datetime(2026, 1, 16))

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Thomas Building 27035.xlsx"), infer_manual=True)

    assert row.values[10:14] == [7, "COMMERCIAL", 3, "CS"]
    assert row.values[19:24] == [42, 70.31, 46, 17, None]


def test_sheet1_thomas_later_knockdown_batch_stays_kd(tmp_path: Path) -> None:
    wb = thomas_sheet1_knockdown_cavity_book(dt.datetime(2026, 3, 25))

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Thomas Building 28499.xlsx"), infer_manual=True)

    assert row.values[11] == "KD"


def test_sheet1_stud_brackets_go_to_w_and_keep_parts_decimals(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Ausmet Job # 28686"
    ws["A11"] = "Material"
    ws["B11"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "OVERALL", "OVERALL ", None, "HINGE ", None, "STRIKER ", "STRIKER ", "DOOR", "STUD"],
        start=1,
    ):
        ws.cell(13, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGE TYPE", "TYPE", "HEIGHT", "CLOSER", "BRACKETS"],
        start=1,
    ):
        ws.cell(14, col).value = value
    door_numbers = ["D.03/SWG.09", "D17/SW1.07", "D.09/SEG.33", "D.10/SE1.06"]
    for offset, door_number in enumerate(door_numbers):
        row_idx = 15 + offset
        values = [
            door_number,
            "SI-12",
            2057,
            1250,
            "DOUBLE",
            "8 (4 EACH SIDE)",
            "100X75X2.5",
            "-",
            "-",
            "-",
            "8 (4 EACH SIDE)",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Piara Waters West PS 28686.xlsx"), infer_manual=True)

    assert row.values[20:23] == [77.76, 32, 32]


def test_sheet1_hinge_jamb_hinge_qty_goes_to_w(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "AUSMET JOB # 28949 REV00"
    ws["A3"] = "Builder"
    ws["B3"] = "Velocity Building"
    ws["A8"] = "Material:"
    ws["B8"] = "1.05mm Zincanneal ZF100"
    for col, value in enumerate([None, None, "OVERALL", "REVEAL", "HAND", "HINGE ", "WELDED"], start=1):
        ws.cell(10, col).value = value
    for col, value in enumerate(["Door #", "PROFILE", "HEIGHT", "WIDTH", None, "QTY", "HINGES"], start=1):
        ws.cell(11, col).value = value
    values = ["RH HINGE JAMB", "A", 2182, "N/A", "N/A", 4, "100X75X2.5"]
    for col, value in enumerate(values, start=1):
        ws.cell(12, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "Velocity Building 28949.xlsx"), infer_manual=True)

    assert row.values[20:23] == [5.72, None, 4]


def test_sheet1_delivery_address_splits_zone_prefix_from_same_cell(tmp_path: Path) -> None:
    cases = [
        ("42b 28 SIGNAL TCE, COCKBURN TCE", "42B", "28 SIGNAL TCE, COCKBURN TCE"),
        ("03C, 62 CLAYTON STREET BELLEVUE", "03C", "62 CLAYTON STREET BELLEVUE"),
    ]

    for raw_address, expected_zone, expected_address in cases:
        wb = sheet1_profile_book()
        ws = wb["Sheet1"]
        ws["A5"] = "Delivery Address"
        ws["B5"] = raw_address
        ws.cell(13, 1).value = 1
        ws.cell(13, 2).value = "Modern"

        row = extract.extract_workbook(save_workbook(wb, tmp_path / f"{expected_zone}.xlsx"), infer_manual=True)

        assert row.values[4] == expected_zone
        assert row.values[5] == expected_address


def test_sheet1_delivery_zone_label_is_used(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws["A8"] = "Delivery Zone:"
    ws["B8"] = "30a"
    values = [1, "Modern", 2060, 923, "RIGHT", 3, "", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-delivery-zone.xlsx"), infer_manual=True)

    assert row.values[4] == "30A"


def test_sheet1_inline_po_label_is_used(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws["H5"] = "PO: 2105.022"
    values = [1, "Modern", 2060, 923, "RIGHT", 3, "", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-inline-po.xlsx"), infer_manual=True)

    assert row.values[1] == "2105.022"


def test_sheet1_material_split_across_adjacent_cells(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws["A10"] = "1.2mm"
    ws["B10"] = "Zinc"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-split-material.xlsx"), infer_manual=True)

    assert row.values[9] == "1.2Z"


def test_sheet1_date_label_is_used_for_delivery_date(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws["A6"] = "Date:"
    ws["B6"] = dt.datetime(2026, 2, 5)
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-date-label.xlsx"), infer_manual=True)

    assert row.values[14:17] == [dt.date(2026, 2, 5), dt.date(2026, 2, 4), "Wednesday"]


def test_standard_worksheet_location_label_can_be_far_right(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    ws["A7"] = "Location:"
    ws["G7"] = "33b"
    values = ["1.2mm Zinc", "", 1, "Modern", "95", 2060, 923, "RIGHT", 3, "HINGE PREP", "S1", 1000]
    for col, value in enumerate(values, start=1):
        ws.cell(10, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-location-far-right.xlsx"), infer_manual=True)

    assert row.values[4] == "33B"


def test_sheet1_hinges_quantity_goes_to_v_bucket(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, "12 + 6"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "SPLIT"]
    assert row.values[19:24] == [2, 5, 4, None, None]


def test_sheet1_csk_dtna_counts_two_parts_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "CSK"
    ws.cell(11, 6).value = "DTNA"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-csk-dtna.xlsx"), infer_manual=True)

    assert row.values[21] == 20
    assert row.values[20] == 20


def test_sheet1_csk_dyna_tube_counts_four_parts_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "CSK"
    ws.cell(11, 6).value = "DYNA AND TUBE"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-csk-dyna-tube.xlsx"), infer_manual=True)

    assert row.values[21] == 40
    assert row.values[20] == 40


def test_main_sheet_csk_dyna_counts_two_parts_per_quantity(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUSMET JOB # 30557 REV00"
    ws["A2"] = "Builder:"
    ws["B2"] = "Australian Fire Door Company"
    ws["A4"] = "Delivery Address:"
    ws["B4"] = "16 Rehill Road, Maddington"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 6, 26)
    ws["A6"] = "Purchase Order:"
    ws["B6"] = "DC 3248"
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(
        [
            "Door #",
            "PROFILE",
            "OVERALL HEIGHT",
            "OVERALL WIDTH",
            "HAND",
            "HINGE QTY",
            "BACKING PLATE TO SUIT",
            "STRIKER TYPE",
            "STRIKER HEIGHT",
            "CSK DYNA",
        ],
        start=1,
    ):
        ws.cell(12, col).value = value
    rows = {
        13: ["D1", "A", 2120, 1010, "LEFT", 4, "100X100X2.5", "S1", 1030, "13 (5 TO EACH JAMB + 3 TO HEAD)"],
        14: ["D2", "A", 2120, 1010, "LEFT", 4, "100X100X2.5", "S1", 1030, "13 (5 TO EACH JAMB + 3 TO HEAD)"],
        15: ["D3", "A", 2110, 1010, "LEFT", 4, "100X100X2.5", "S1", 1030, "13 (5 TO EACH JAMB + 3 TO HEAD)"],
        16: ["D4", "B", 2082, 997, "RIGHT", 4, "100X100X2.5", "S1", 1030, "13 (5 TO EACH JAMB + 3 TO HEAD)"],
        17: ["D5", "B", 2074, 996, "RIGHT", 4, "100X100X2.5", "S1", 1030, "13 (5 TO EACH JAMB + 3 TO HEAD)"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "csk-dyna-main.xlsx"), infer_manual=True)

    assert row.values[21] == 135


def test_sheet1_tradition_dyna_counts_one_part_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "TRADITION"
    ws.cell(11, 6).value = "DYNA"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-tradition-dyna.xlsx"), infer_manual=True)

    assert row.values[21] == 10
    assert row.values[20] == 10


def test_sheet1_hinge_plates_quantity_goes_to_w_bucket(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(12, 7).value = "HINGE PLATES"
    values = [2, "Split A + B", 2060, 923, "RIGHT", 3, "Suit 100x75x2.5", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-hinge-plates.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "SPLIT"]
    assert row.values[19:24] == [4, 11, 2, 6, None]


def test_sheet1_backing_plate_quantity_goes_to_w_bucket(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 26944 REV00"
    for col, value in enumerate(
        ["Door #", "Frame ", "Wall ", "REVEAL", "REVEAL", None, "DOOR", "HINGE ", "BACKING PLATE", "STRIKER ", "STRIKER ", "DYNA"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        [None, "Type", "Type", "HEIGHT", "WIDTH", "HAND", "TYPE", "QTY", "TO SUIT", "TYPE", "HEIGHT", "BOLT"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["D1/1.19", "SD01", "SP-01", 2360, 923, "LEFT", "SC = 40mm", 4, "100X75X2.5", "-", "-", "-"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-sheet-backing-plate.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 6, 0, 4, None]


def test_trad_dyna_backing_plate_is_not_counted_by_default(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 27380 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(
        [None, None, "OVERALL", "OVERALL", None, "HINGE ", "BACKING PLATE", "STRIKER ", "STRIKER ", "DYNA", "TRAD"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "BOLT", "DYNA"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["A809.1", "H", 2090, 1015, "RIGHT", 4, "100X100X2.5", "S1", 1020, None, "13 (5 EACH JAMB + 3 IN HEAD)"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "trad-dyna-backing-plate.xlsx"), infer_manual=True)

    assert row.values[21:23] == [14, 4]


def test_offset_brick_ties_and_wire_ties_count_two_each_in_v_bucket(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 27434 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "REVEAL", "REVEAL", None, "HINGE ", "BACKING PLATE", "MORTAR GUARD", "STRIKER ", "STRIKER ", "DYNA", "OFFSET BRICK", "WIRE"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TO HINGE PLATE", "TYPE", "HEIGHT", "BOLT", "TIES", "TIES"],
        start=1,
    ):
        ws.cell(12, col).value = value
    rows = {
        13: ["A1", "A", 2060, 1030, "LEFT", 4, "100x100x2.5", "YES", "ES2100", 1030, None, "8-BTOFF", "NO"],
        14: ["A2", "A", 2060, 1030, "LEFT", 4, "100x100x2.5", "YES", "-", 1030, None, "NO", "1 BAG"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "brick-ties.xlsx"), infer_manual=True)

    assert row.values[21] == 5


def test_numeric_offset_brick_ties_are_not_counted_as_hardware(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 28070 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "REVEAL", "REVEAL", None, "HINGE ", "BACKING PLATE", "MORTAR GUARD", "STRIKER ", "STRIKER ", "DYNA", "OFFSET BRICK"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TO HINGE PLATE", "TYPE", "HEIGHT", "BOLT", "TIES"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["A1", "A", 2060, 1030, "LEFT", 4, "100x100x2.5", "YES", "S1", 1030, None, "8"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "numeric-offset-brick-ties.xlsx"), infer_manual=True)

    assert row.values[21] == 1


def test_wire_ties_are_not_counted_when_trad_dyna_column_exists(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 26974 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "REVEAL", "REVEAL", None, "HINGE ", "MORTAR GUARD", "STRIKER ", "STRIKER ", "DYNA", "TRAD", "WIRE"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO HINGE PLATE", "TYPE", "HEIGHT", "BOLT", "DYNA", "TIES"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["A1", "A", 2060, 1030, "LEFT", 4, "YES", "S1", 1020, None, "30 (12 EACH JAMB +6 IN HEAD)", "1 BAG"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "trad-dyna-wire-ties.xlsx"), infer_manual=True)

    assert row.values[21] == 31


def test_trad_dyna_es2100_uses_each_jamb_qty_for_striker(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 30534 REV00"
    ws["A2"] = "Builder:"
    ws["B2"] = "Australian Fire Door Company"
    ws["A4"] = "Delivery Address:"
    ws["B4"] = "03c AFDC, Chalkley Place Bayswater"
    ws["A5"] = "Delivery Date:"
    ws["B5"] = dt.date(2026, 6, 26)
    ws["A6"] = "Purchase Order:"
    ws["B6"] = "SO 7421"
    ws["A9"] = "Material:"
    ws["B9"] = "1.6mm Zinc"
    for col, value in enumerate(
        [
            "Door #",
            "PROFILE",
            "OVERALL HEIGHT",
            "OVERALL WIDTH",
            "HAND",
            "HINGE QTY",
            "BACKING PLATE TO SUIT",
            "STRIKER TYPE",
            "STRIKER HEIGHT",
            "TRAD DYNA",
        ],
        start=1,
    ):
        ws.cell(12, col).value = value
    rows = {
        13: ["SO_7421/1.18 STAIR 4", "A", 2100, 1005, "RIGHT", 4, "100X100X2.5", "ES2100", 1020, "13 (5 EACH JAMB + 3 IN HEAD)"],
        14: ["SO_7421/01.14A STAIR 1", "C", 2100, 1005, "LEFT", 4, "100X100X2.5", "ES2100", 1020, "13 (5 EACH JAMB + 3 IN HEAD)"],
        15: ["SO_7421/01.14B FIRE CNTR", "D", 2100, 1005, "LEFT", 4, "100X100X2.5", "ES2100", 1020, "13 (5 EACH JAMB + 3 IN HEAD)"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "trad-dyna-es2100.xlsx"), infer_manual=True)

    assert row.values[21] == 54


def test_trad_dyna_backing_plate_is_not_counted_in_brick_tables(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 26974 REV00"
    ws["A9"] = "Material:"
    ws["B9"] = "1.2mm Zinc"
    for col, value in enumerate(
        [None, None, "REVEAL", "REVEAL", None, "HINGE ", "BACKING PLATE", "MORTAR GUARD", "STRIKER ", "STRIKER ", "DYNA", "TRAD", "WIRE"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TO HINGE PLATE", "TYPE", "HEIGHT", "BOLT", "DYNA", "TIES"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["A1", "A", 2060, 1030, "LEFT", 4, "100x100x2.5", "YES", "S1", 1020, None, "30 (12 EACH JAMB +6 IN HEAD)", "NO"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "brick-trad-dyna-backing.xlsx"), infer_manual=True)

    assert row.values[21] == 31


def test_main_sheet_striker_height_is_not_counted_as_hardware_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 27033 REV00"
    for col, value in enumerate(
        ["Door #", "Frame ", "Wall ", "REVEAL", "REVEAL", None, "DOOR", "HINGE ", "BACKING PLATE", "STRIKER ", "STRIKER ", "DYNA"],
        start=1,
    ):
        ws.cell(11, col).value = value
    for col, value in enumerate(
        [None, "Type", "Type", "HEIGHT", "WIDTH", "HAND", "TYPE", "QTY", "TO SUIT", "TYPE", "HEIGHT", "BOLT"],
        start=1,
    ):
        ws.cell(12, col).value = value
    values = ["D1/1.03", "SD04", "SP-01 + LINING", 2360, 925, "RIGHT", "SC = 40mm", 4, "100X75X2.5", "MORTICE LOCK (S1)", 1032, None]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-sheet-striker-height.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 7, 1, 4, None]


def test_sheet1_profile_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, "12 + 6"]
    for row_idx in (13, 14):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[14].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-sheet1.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "SPLIT"]
    assert row.values[19:24] == [2, 5, 4, None, None]


def test_sheet1_profile_code_maps_to_commercial_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "144B/O", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-commercial-code.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_sheet1_cavity_slider_contributes_fourteen_mitres_each(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 1).value = "Cavity Sliders"
    values = [2, "Modern", 2060, 700, "", "", "", "", "", ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-cavity-slider.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "CS"]
    assert row.values[19] == 28


def test_sheet1_numeric_letter_profile_code_maps_to_commercial_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "150B", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-commercial-number-letter.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_sheet1_kd_abbreviation_maps_to_kd_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "144mm B/O KD - Screw Fix Mitre", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-kd-code.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]
    assert row.values[19:24] == [0, 7, 4, None, 7]


def test_sheet1_head_only_replacement_rows_do_not_count_as_kd_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    rows = {
        13: [1, "125mm B/O MODERN KNOCK DOWN-SCREW FIXED MITRE", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""],
        14: ["REPLACEMENT HEAD # 1", "112mm B/O MODERN KNOCK DOWN-SCREW FIXED MITRE HEAD ONLY", "", 723, "LEFT", "", "", "", "", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-head-only.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]


def test_standard_worksheet_head_only_order_still_counts_as_kd_goods(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Modern Knockdown - screw fix HEAD ONLY",
        "114",
        0,
        1315,
        "",
        "",
        "",
        "",
        "",
        "NO",
        "NO",
        "NO",
        "",
        "",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "head-only-order.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]


def test_sheet1_service_part_maps_to_part_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [50, "Service Part - Hinge Plate Suit Fire Door", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-service-part.xlsx"), infer_manual=True)

    assert row.values[10:12] == [50, "PART"]


def main_sheet_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    wb.create_sheet("Profiles")
    ws["A1"] = "Job No 30354"
    ws["B2"] = "Fire Door Maintenance"
    headers = ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "GUARDS", "TYPE", "HEIGHT", "BOLT", "BRICK TIES"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    return wb


def test_main_sheet_commercial_widths_use_reveal_1220_over_size_threshold(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    rows = {
        13: ["D.4", "A", 2410, 1920, "DOUBLE", "8 (4 EACH SIDE)", "100X100X2.5", "YES", "-", "-", "-", 10],
        14: ["D.3", "A", 2110, 1030, "RIGHT", 4, "100X100X2.5", "YES", "S1", 1020, "-", 8],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profile.xlsx"), infer_manual=True)

    assert row.values[7] == 1
    assert row.values[10:12] == [2, "COMMERCIAL"]
    assert row.values[19:24] == [3, 18, 1, 12, None]


def test_main_sheet_split_frame_numbered_parts_count_once(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Quantity", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGE PLATES", "TYPE", "HEIGHT", "BRICK TIES"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    rows = {
        13: [4, "Split Frame # 1", 2088, 884, "RIGHT", 3, "SUIT 100X75X2.5", "S1", 1039, "8 pieces"],
        14: [4, "Split Frame Frame # 2", 2091, 890, "-", "-", "-", "-", "-", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-split-frame-parts.xlsx"), infer_manual=True)

    assert row.values[10:12] == [4, "SPLIT"]


def test_main_sheet_stillages_note_overrides_width_count(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    ws["A20"] = "PLEASE PACK ONTO A CUSTOM STILLAGE"
    rows = {
        13: ["D.4", "A", 2410, 1920, "DOUBLE", "8 (4 EACH SIDE)", "100X100X2.5", "YES", "-", "-", "-", 10],
        14: ["D.3", "A", 2110, 1030, "RIGHT", 4, "100X100X2.5", "YES", "S1", 1020, "-", 8],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-stillages.xlsx"), infer_manual=True)

    assert row.values[7] == "stillages"


def test_main_sheet_concealed_cavity_height_over_threshold_counts_oversize(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    rows = {
        13: ["1", "114B/O CONCEALED", 2520, 1180, 1230, "4 TO CLOSING JAMB", "", "BRIO", "NO", "NO", "", ""],
        14: ["1", "114B/O CONCEALED", 2660, 930, 980, "4 TO CLOSING JAMB", "", "BRIO", "NO", "NO", "", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-concealed-height.xlsx"), infer_manual=True)

    assert row.values[7] == 2


def test_main_sheet_offset_hand_column_still_counts_oversize_width(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = [
        "Door #",
        "(PER A_6020 & A_6021)",
        "PROFILE",
        "THICKNESS",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "QTY",
        "TO SUIT",
        "TYPE",
        "HEIGHT",
        "BOLT",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "ON.LG035B",
        "13",
        "CUSTOM",
        "40mm",
        2400,
        1856,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-offset-hand.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_main_sheet_old_overall_template_counts_double_width_only(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    ws["A1"] = "Job No 27256"
    header_rows = {
        11: ["", "", "OVERALL", "OVERALL", "DOOR", "", "", ""],
        12: ["Door #", "PROFILE", "HEIGHT", "HEIGHT", "THICKNESS", "HAND", "QTY", "TO SUIT"],
    }
    detail_rows = {
        13: ["1", "A", 2410, 1946, 40, "DOUBLE", 1, ""],
        14: ["2", "A", 2110, 1026, 40, "RIGHT", 1, ""],
        15: ["3", "A", 2110, 1212, 40, "LEFT", 1, ""],
    }
    for row_idx, values in {**header_rows, **detail_rows}.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-old-overall.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_over_size_override_applies_by_job_number(tmp_path: Path) -> None:
    original_rules = extract.RULES
    try:
        extract.RULES = extract.RuleSet(over_size_overrides={"28225": "stillages"})
        wb = main_sheet_book()
        ws = wb["Main Sheet"]
        ws["A1"] = "Job No 28225"

        row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-override.xlsx"), infer_manual=True)

        assert row.values[7] == "stillages"
    finally:
        extract.RULES = original_rules


def test_main_sheet_profileless_table_extracts_commercial_hardware(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "MAIL ROOM",
        "SPLIT 85-125B/O ",
        "40mm",
        2060,
        923,
        "RIGHT",
        4,
        "100X75X2.5",
        "MORTICE LOCK (S1)",
        1032,
        "12 + 6",
        "NOT REQUIRED",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 5, 1, 4, None]


def test_main_sheet_profileless_width_over_threshold_counts_oversize(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "D2.01/S038.1",
        "DMF02",
        "DL01.M = 40mm Thick",
        2380,
        1246,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless-oversize.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_main_sheet_profileless_hand_header_adds_commercial_double_mitre(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "D2.01/S038.1",
        "DMF02",
        "DL01.M = 40mm Thick",
        2380,
        1246,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless-double.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19] == 2


def test_main_sheet_double_action_commercial_adds_half_mitre_each(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = [
        "Door #",
        "(PER A_6020 & A_6021)",
        "PROFILE",
        "THICKNESS",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "DOUBLE ACTION BOXES",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    for row_idx in (13, 14):
        values = [f"ON.L{row_idx}", "08A", "CUSTOM D/A", "40mm", 2260, 2272, "DOUBLE ACTION", "2 - WELDED IN TO HEAD OF FRAME"]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-double-action.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "COMMERCIAL"]
    assert row.values[19] == 3


def test_main_sheet_profileless_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "MAIL ROOM",
        "SPLIT 85-125B/O ",
        "40mm",
        2060,
        923,
        "RIGHT",
        4,
        "100X75X2.5",
        "MORTICE LOCK (S1)",
        1032,
        "12 + 6",
        "NOT REQUIRED",
    ]
    for row_idx in (13, 14):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[14].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-main-profileless.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 5, 1, 4, None]


def test_main_sheet_profile_hash_header_uses_one_product_when_qty_is_hardware(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    for col in range(1, 13):
        ws.cell(12, col).value = None
    ws.cell(15, 7).value = "HINGE"
    headers = [
        "Door #",
        "PROFILE #",
        "TYPE",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "QTY",
        "SIZE",
        "TYPE",
        "HEIGHT",
        "Switch",
        "Switch Hole",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(16, col).value = value
    values = ["00-059-D1", "CUSTOM D1", "DLFRCL", 2410, 1062, "RIGHT", 5, "100X100X3.2", "S1", 1032, "1 IN HEAD", "NO"]
    for col, value in enumerate(values, start=1):
        ws.cell(17, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profile-hash.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_main_sheet_split_part_only_counts_one_mitre_each(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [
        8,
        "CUSTOM SPLIT PART B ONLY (30MM ARCHITRAVE)",
        2060,
        910,
        "RIGHT",
        8,
        "100X75X2.5",
        "S1",
        1000,
        "-",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "split-part-b-only.xlsx"), infer_manual=True)

    assert row.values[10:12] == [8, "SPLIT"]
    assert row.values[19] == 8


def test_nonstandard_worksheet_door_qty_rows_extract_goods_qty(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["A1"] = "Ausmet Job # 29463"
    ws["C2"] = "Carnarvon Timber & Hardware"
    ws["C5"] = "2026-05-13"
    headers = [
        "Material",
        "Door #",
        "B/O",
        "Qty",
        "WALL TYPE",
        "FRAME TYPE",
        "DOOR THICKNESS",
        "REVEAL HEIGHT",
        "REVEAL WIDTH",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(10, col).value = value
    rows = {
        11: ["1.05mm Galv", "", "", 0, "", "", "", "", ""],
        12: ["1.05mm Galv", "AC BOX SPLIT A & B", "", 1, "", "", "", 560, 770],
        13: ["1.05mm Galv", "AC BOX SPLIT A & B", "", 1, "", "", "", 575, 700],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "nonstandard-split.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "SPLIT"]
    assert row.values[19] == 4


def test_door_skins_capping_rows_do_not_fallback_to_commercial_goods(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 29322"
    ws["B2"] = "Australian Fire Door Company"
    headers = ["MATERIAL", "PROFILE", "QUANTITY", "WIDTH", "LENGTH", "BOLT", ""]
    for col, value in enumerate(headers, start=1):
        ws.cell(11, col).value = value
    rows = {
        12: ["0.55mm Deep Ocean", "A", 2, 845, 2030, "", "FLAT SHEET"],
        13: ["0.55mm Deep Ocean", "B", 2, "", 2100, "", "CAPPING"],
        14: ["0.55mm Deep Ocean", "B", 2, "", 1000, "", "CAPPING"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "AFDC Door Skins.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "DS"]


def test_door_skins_filename_with_profile_code_maps_to_ds_only(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 30471 REV00"
    ws["B2"] = "Danze Mining"
    ws["B4"] = "2 Volcanic Lp, Wangara"
    ws["B5"] = dt.datetime(2026, 6, 24)
    ws["B6"] = "02-6365-03"
    ws["B7"] = "57d"
    ws["B10"] = "1.2mm"
    ws["C10"] = "Zinc"
    headers = ["Quantity", "PROFILE", "HEIGHT", "WIDTH", "FOUR SIDED"]
    for col, value in enumerate(headers, start=1):
        ws.cell(13, col).value = value
    values = [4, "A", 368, 268, "YES"]
    for col, value in enumerate(values, start=1):
        ws.cell(14, col).value = value

    row = extract.extract_workbook(
        save_workbook(wb, tmp_path / "Danze Mining 30471 Door Skins 2mm and 1.05 + Window Handle Blank.xlsx"),
        infer_manual=True,
    )

    assert row.values[10:12] == [4, "DS"]
    assert row.values[19:24] == [None, None, 0, None, None]


def test_trad_dyna_split_profile_counts_as_commercial_goods(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    rows = {
        13: ["D8", "A", 2075, 1011, "LEFT", 4, "100x100x2.5", "S1", 1030, "", ""],
        20: ["D7", "125BO SPLIT", 2060, 910, "LEFT", 4, "100X100X2.5", "S1", 1000, "-", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "AFDC Trad Dyna Single Rebate.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "COMMERCIAL"]
