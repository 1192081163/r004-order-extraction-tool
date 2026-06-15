#!/usr/bin/env python3
"""
Extract job-tracking rows from Ausmet order workbooks.

The script reads order .xlsx/.xlsm files and writes a CSV shaped like the
tracking workbook's A-X columns.

By default it only fills fields that are directly supported by the order file.
Columns that are normally scheduled, estimated, or manually summarized are left
blank. Use --infer-manual only when you explicitly want best-effort estimates
for supported manual columns.

Usage:
    python3 extract_job_track.py --input-dir . --output extracted_job_rows.csv
    python3 extract_job_track.py --input-dir . --output extracted_job_rows.csv --infer-manual
    python3 extract_job_track.py --input-dir . --output extracted_job_rows.csv --compare-total "total.xlsx"
    python3 extract_job_track.py --input-dir . --output extracted_job_rows.csv --audit-output audit.csv
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


TRACK_HEADERS = [
    "D DATE",
    "PO NUMBER",
    "BUILDER",
    "Urgent",
    "",
    "Deliver Address",
    "#",
    "Over \nSize",
    "",
    "Material",
    "QTY",
    "Goods1",
    "QTY",
    "Goods2",
    "Ideal\n D date",
    "Estimate\n C Date",
    "",
    "MITRE",
    "Parts",
    "MITRE",
    "Parts",
    "hinge/striker/stud/dynabolt/2110/sill",
    "hinge holder/3751/WS7/mib",
    "door closer/kd",
]

COLUMN_LETTERS = [chr(ord("A") + idx) for idx in range(len(TRACK_HEADERS))]

MANUAL_OR_SCHEDULED_COLUMNS = {
    7,   # H Over Size
    15,  # P Estimate C Date
    16,  # Q weekday
    17,  # R summary MITRE
    18,  # S summary Parts
    19,  # T per-job MITRE
    20,  # U calculated/display Parts
    21,  # V hardware bucket
    22,  # W hardware bucket
    23,  # X hardware bucket
}

OVER_SIZE_WIDTH_THRESHOLD = 1024

COLORBOND_COLOUR_MARKERS = {
    "basalt",
    "classic cream",
    "cove",
    "deep ocean",
    "dune",
    "gully",
    "ironstone",
    "jasper",
    "loft",
    "mangrove",
    "manor red",
    "monument",
    "night sky",
    "pale eucalypt",
    "paperbark",
    "shale grey",
    "southerly",
    "surfmist",
    "terrain",
    "wallaby",
    "wilderness",
    "windspray",
    "woodland grey",
}

CHINA_2026_HOLIDAYS = {
    dt.date(2026, 1, 1),
    dt.date(2026, 1, 2),
    dt.date(2026, 1, 3),
    dt.date(2026, 2, 15),
    dt.date(2026, 2, 16),
    dt.date(2026, 2, 17),
    dt.date(2026, 2, 18),
    dt.date(2026, 2, 19),
    dt.date(2026, 2, 20),
    dt.date(2026, 2, 21),
    dt.date(2026, 2, 22),
    dt.date(2026, 2, 23),
    dt.date(2026, 4, 4),
    dt.date(2026, 4, 5),
    dt.date(2026, 4, 6),
    dt.date(2026, 5, 1),
    dt.date(2026, 5, 2),
    dt.date(2026, 5, 3),
    dt.date(2026, 5, 4),
    dt.date(2026, 5, 5),
    dt.date(2026, 6, 19),
    dt.date(2026, 6, 20),
    dt.date(2026, 6, 21),
    dt.date(2026, 9, 25),
    dt.date(2026, 9, 26),
    dt.date(2026, 9, 27),
    dt.date(2026, 10, 1),
    dt.date(2026, 10, 2),
    dt.date(2026, 10, 3),
    dt.date(2026, 10, 4),
    dt.date(2026, 10, 5),
    dt.date(2026, 10, 6),
    dt.date(2026, 10, 7),
}

CHINA_2026_ADJUSTED_WORKDAYS = {
    dt.date(2026, 1, 4),
    dt.date(2026, 2, 14),
    dt.date(2026, 2, 28),
    dt.date(2026, 5, 9),
    dt.date(2026, 9, 20),
    dt.date(2026, 10, 10),
}


BUILDER_ALIASES = {
    "a&m construction group": "A&M",
    "aart homes": "AART",
    "activa homes group pty ltd": "ACTIVA",
    "ace wa construction pty ltd": "ACE",
    "alita construction": "ALITA",
    "apex building (aus) pty ltd": "APEX",
    "austern building supplies pty ltd": "AUSTERN",
    "aw design & build pty ltd": "AW",
    "beaumonde homes": "Beaumonde",
    "beyond residential": "BEYOND",
    "csr building products ltd": "CSR",
    "bunnings group limited": "Bunnings",
    "building development group constructions pty ltd": "BDGC",
    "built": "BULIT",
    "cash sale": "CASH SALE",
    "cash sale - loris moriconi (abn staff member)": "CASH SALE",
    "celebration homes": "CELEBRATION",
    "coastal design & construction pty ltd": "COASTAL",
    "c u building group": "CU",
    "customised projects": "CUSTOMISED",
    "dale alcock homes": "DALE ALCOCK",
    "danze mining": "DANZE",
    "danze mining & building products": "DANZE",
    "dasco building group pty ltd": "DASCO",
    "distinct homes pty ltd": "DISTINCTIVE",
    "distinctive homes wa pty ltd": "DISTINCTIVE",
    "direct homes wa": "DIRECT",
    "dynamic steelform": "DYNAMIC",
    "edge construction": "EDGE",
    "fire door maintenance": "FDM",
    "fire door mainenance": "FDM",
    "fratelli homes (wa) pty ltd": "FRATELLI",
    "geared construction": "GEARED",
    "giorgi architects + builders (building corporation": "GIORGI",
    "gvm solutions ptuy ltd": "GVM",
    "australian fire door company": "AFDC",
    "evoke living homes": "EVOKE",
    "homebuyers centre": "HOMEBUYERS",
    "imagine building wa pty ltd": "IMAGINE",
    "indoz homez pty ltd": "INDOZ",
    "insite residential": "INSITE",
    "ionic projects pty ltd": "IONIC",
    "karlin supplies": "KARLIN",
    "la vida australia pty ltd": "La Vida",
    "lee contracting group pty ltd": "LEE",
    "leigh homes pty ltd": "LEIGH",
    "lock up security & doors (polon pty ltd t/a)": "LOCK UP",
    "longhua international pty ltd": "LONGHUA",
    "louis homes pty ltd": "LOUIS",
    "makin homes": "MAKIN",
    "m&b sales": "M&B",
    "marshall homes pty ltd": "MARSHALL",
    "mecca constructions pty ltd": "MECCA",
    "midstream hardware (ragra pty ltd t/a)": "MIDSTREAM",
    "modular wa": "MODULAR",
    "mvg construction pty ltd": "MVG",
    "new era homes australia": "NEW ERA",
    "nexus contruction co": "NEXUS",
    "nu-style living": "NUSTYLE",
    "novus homes (antonelli investments p/l trading as)": "Novus",
    "oceanic custom homes": "OCEANIC",
    "one stop doors": "OSD",
    "oz home building": "OZ",
    "papalia building & design pty ltd": "PAPALIA",
    "planet building products pty ltd": "PLANET",
    "prestige homes wa pty ltd": "Prestige",
    "prima homes": "PRIMA",
    "prime projects construction p/l": "PRIME",
    "project building supplies ( plasterboard projects)": "PROJECT",
    "project building supplies south west": "PROJECT",
    "rg construct pty ltd": "RG",
    "ross north homes the challengerhomes unit trust": "ROSS NORTH",
    "ryza homes (s&p glossop t/a)": "RYZA",
    "select living": "SELECT",
    "spence doors": "SPENCE",
    "superior homes": "SUPERIOR",
    "thomas building": "THOMAS",
    "the homesmith group": "HOMESMITH",
    "tj payne developments": "TJ",
    "tobia constructions": "TOBIA",
    "trio home builders pty ltd": "TRIO",
    "viva developments pty ltd": "VIVA",
    "vm building": "VM",
    "westwood homes": "Westwood",
    "willing build": "WILLING",
    "zz designer homes": "ZZ",
}


PROFILE_ALIASES = {
    "a": "COMMERCIAL",
    "b": "COMMERCIAL",
    "c": "COMMERCIAL",
    "custom": "COMMERCIAL",
    "d": "COMMERCIAL",
    "h": "COMMERCIAL",
    "j": "COMMERCIAL",
    "w": "COMMERCIAL",
}


@dataclass
class ExtractedRow:
    values: list[Any] = field(default_factory=lambda: [None] * len(TRACK_HEADERS))
    notes: list[str] = field(default_factory=list)
    manual_check: list[str] = field(default_factory=list)
    source_file: str = ""

    def add_manual_check(self, message: str) -> None:
        if message not in self.manual_check:
            self.manual_check.append(message)


@dataclass(frozen=True)
class RuleSet:
    builder_aliases: dict[str, str] = field(default_factory=dict)
    goods_ignore_patterns: tuple[str, ...] = ()
    holidays: frozenset[dt.date] = frozenset()
    adjusted_workdays: frozenset[dt.date] = frozenset()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = text.replace("*", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def rule_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def load_rules(rules_dir: Path) -> RuleSet:
    builder_aliases: dict[str, str] = {}
    goods_ignore_patterns: list[str] = []
    holidays = set(CHINA_2026_HOLIDAYS)
    adjusted_workdays = set(CHINA_2026_ADJUSTED_WORKDAYS)

    builder_path = rules_dir / "builder_aliases.csv"
    if builder_path.exists():
        with builder_path.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                source = rule_key(row.get("source"))
                builder = clean_text(row.get("builder"))
                if source and builder:
                    builder_aliases[source] = builder

    goods_path = rules_dir / "goods_ignore_patterns.csv"
    if goods_path.exists():
        with goods_path.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                pattern = rule_key(row.get("pattern"))
                if pattern:
                    goods_ignore_patterns.append(pattern)

    calendar_path = rules_dir / "china_workdays_2026.csv"
    if calendar_path.exists():
        holidays.clear()
        adjusted_workdays.clear()
        with calendar_path.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                raw_date = clean_text(row.get("date"))
                raw_type = rule_key(row.get("type"))
                if not raw_date:
                    continue
                day = dt.datetime.strptime(raw_date, "%Y-%m-%d").date()
                if raw_type == "holiday":
                    holidays.add(day)
                elif raw_type == "workday":
                    adjusted_workdays.add(day)

    return RuleSet(
        builder_aliases=builder_aliases,
        goods_ignore_patterns=tuple(goods_ignore_patterns),
        holidays=frozenset(holidays),
        adjusted_workdays=frozenset(adjusted_workdays),
    )


RULES = load_rules(Path(__file__).with_name("rules"))


def normalize_delivery_address(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if re.sub(r"[^A-Za-z0-9]+", "", text).casefold() == "pickup":
        return "PICK UP"
    text = re.sub(r"\bPICKUP\b", "PICK UP", text, flags=re.IGNORECASE)
    return text


def normalize_po_number(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"\s*\[[^\]]+\]\s*$", "", text).strip()
    if re.fullmatch(r"\d+", text):
        return text.lstrip("0") or "0"
    return text


def sheet_has_door_skin_profile(ws: Any) -> bool:
    for row_idx in range(1, min(ws.max_row, 180) + 1):
        for col_idx in range(1, ws.max_column + 1):
            text = clean_text(ws.cell(row_idx, col_idx).value).upper()
            if "SKIN" in text:
                return True
    return False


def normalize_danze_door_skin_po_number(value: Any, builder: Any, ws: Any) -> str | None:
    text = normalize_po_number(value)
    if clean_text(builder).upper() != "DANZE":
        return text
    if not text or not sheet_has_door_skin_profile(ws):
        return text
    match = re.fullmatch(r"\d{2}-(\d+)-\d{2}", text)
    if not match:
        return text
    return match.group(1).lstrip("0") or "0"


def number_or_none(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    num = float(match.group(0))
    return int(num) if num.is_integer() else num


def excel_display_int(value: Any) -> int | None:
    """Match Excel's 0-decimal display for positive/negative numeric values."""
    if value is None or value == "":
        return None
    dec = Decimal(str(value))
    return int(dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def over_size_marker_from_text(value: Any) -> str | None:
    upper = clean_text(value).upper()
    if not upper:
        return None
    if "GLUT" in upper:
        return "glut"
    if "PALLET" in upper:
        return "pallet"
    if "STILLAGE" in upper:
        return "stillages"
    return None


def worksheet_over_size_marker(ws: Any) -> str | None:
    markers: list[str] = []
    for row_idx in range(1, ws.max_row + 1):
        line = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 24) + 1))
        marker = over_size_marker_from_text(line)
        if marker:
            markers.append(marker)
    for marker in ("glut", "pallet", "stillages"):
        if marker in markers:
            return marker
    return None


def workbook_over_size_marker(wb: Any) -> str | None:
    markers = [worksheet_over_size_marker(ws) for ws in wb.worksheets]
    for marker in ("glut", "pallet", "stillages"):
        if marker in markers:
            return marker
    return None


def write_over_size(row: ExtractedRow, marker: str | None, qty: float = 0.0) -> None:
    # Tracking H is mostly the count of rows with width > 1024; packaging notes use text.
    if marker in {"glut", "pallet"}:
        row.values[7] = marker
    elif qty:
        row.values[7] = excel_display_int(qty)
    elif marker:
        row.values[7] = marker


def width_value(value: Any) -> float:
    num = number_or_none(value)
    return float(num or 0)


def worksheet_over_size_qty(ws: Any) -> float:
    qty_total = 0.0
    for idx in worksheet_rows(ws):
        if clean_text(ws[f"A{idx}"].value).lower() == "other":
            continue
        qty = number_or_none(ws[f"C{idx}"].value) or 0
        if qty <= 0:
            continue
        # G is reveal width in the standard table. In cavity rows I is overall width;
        # in standard rows I is hinge qty, so including it does not affect the threshold.
        max_width = max(width_value(ws[f"G{idx}"].value), width_value(ws[f"I{idx}"].value))
        if max_width > OVER_SIZE_WIDTH_THRESHOLD:
            qty_total += float(qty)
    return qty_total


def sheet1_width_columns(ws: Any, header_row: int) -> list[int]:
    return [
        col_idx
        for col_idx in range(1, ws.max_column + 1)
        if "WIDTH" in sheet1_header_text(ws, header_row, col_idx)
    ]


def row_looks_cavity_table_header(ws: Any, row_idx: int) -> bool:
    line = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 24) + 1)).upper()
    return "OVERALL CAVITY" in line or ("POCKET" in line and "SLIDER" in line)


def cavity_without_profile_over_size_qty(ws: Any) -> float:
    qty_total = 0.0
    for header_row in range(1, min(ws.max_row, 180) + 1):
        if not row_looks_cavity_table_header(ws, header_row):
            continue
        header_context = (
            " ".join(clean_text(ws.cell(header_row, col).value) for col in range(1, min(ws.max_column, 24) + 1))
            + " "
            + " ".join(clean_text(ws.cell(header_row + 1, col).value) for col in range(1, min(ws.max_column, 24) + 1))
        ).upper()
        if "PROFILE" in header_context:
            continue
        width_cols = [
            col_idx
            for col_idx in range(1, ws.max_column + 1)
            if "WIDTH"
            in (
                clean_text(ws.cell(header_row, col_idx).value)
                + " "
                + clean_text(ws.cell(header_row + 1, col_idx).value)
            ).upper()
        ]
        for row_idx in range(header_row + 2, ws.max_row + 1):
            first = clean_text(ws.cell(row_idx, 1).value)
            if not first or first.upper().startswith(("DOOR", "MATERIAL")):
                break
            max_width = max([width_value(ws.cell(row_idx, col).value) for col in width_cols] or [0])
            if max_width > OVER_SIZE_WIDTH_THRESHOLD:
                qty_total += 1
    return qty_total


def profile_tables_over_size_qty(ws: Any) -> float:
    qty_total = 0.0
    headers = find_sheet1_profile_headers(ws)
    for header_index, (header_row, profile_col) in enumerate(headers):
        next_header_row = headers[header_index + 1][0] if header_index + 1 < len(headers) else ws.max_row + 1
        product_qty_col = sheet1_product_qty_column(ws, header_row)
        width_cols = sheet1_width_columns(ws, header_row)
        for idx in range(header_row + 1, next_header_row):
            if row_looks_cavity_table_header(ws, idx):
                break
            profile = ws.cell(idx, profile_col).value
            if not has_value(profile):
                continue
            if clean_text(ws.cell(idx, 1).value).lower().startswith("material"):
                continue
            if material_code(profile):
                continue
            product_qty = sheet1_product_qty(ws, idx, profile_col, product_qty_col)
            if product_qty <= 0:
                continue
            max_width = max([width_value(ws.cell(idx, col).value) for col in width_cols] or [0])
            if max_width > OVER_SIZE_WIDTH_THRESHOLD:
                qty_total += float(product_qty)
    return qty_total + cavity_without_profile_over_size_qty(ws)


def main_width_columns(ws: Any) -> list[int]:
    columns: list[int] = []
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        for col_idx in range(1, ws.max_column + 1):
            context = " ".join(
                clean_text(ws.cell(idx, col_idx).value)
                for idx in range(max(1, row_idx - 1), min(ws.max_row, row_idx + 1) + 1)
            ).upper()
            if "WIDTH" in context and col_idx not in columns:
                columns.append(col_idx)
    return columns


def main_fallback_over_size_qty(ws: Any) -> float:
    width_cols = main_width_columns(ws)
    qty_total = 0.0
    for idx in iter_main_detail_rows(ws):
        max_width = max([width_value(ws.cell(idx, col).value) for col in width_cols] or [0])
        if max_width > OVER_SIZE_WIDTH_THRESHOLD:
            qty_total += 1
    return qty_total + cavity_without_profile_over_size_qty(ws)


def parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_sheet1_delivery_date(value: Any, row: ExtractedRow) -> dt.date | None:
    parsed = parse_date(value)
    if parsed is not None:
        return parsed
    text = clean_text(value)
    if text:
        row.add_manual_check(f"delivery date not parsed: {text}")
    return None


def is_china_workday(day: dt.date) -> bool:
    if day in RULES.adjusted_workdays:
        return True
    if day in RULES.holidays:
        return False
    return day.weekday() < 5


def previous_business_day(day: dt.date | None) -> dt.date | None:
    if day is None:
        return None
    current = day - dt.timedelta(days=1)
    while not is_china_workday(current):
        current -= dt.timedelta(days=1)
    return current


def format_date(value: dt.date | dt.datetime | None) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        value = value.date()
    return value


def normalize_builder(builder: str, row: ExtractedRow) -> str:
    key = clean_text(builder).lower()
    builder_aliases = {**BUILDER_ALIASES, **RULES.builder_aliases}
    if key in builder_aliases:
        return builder_aliases[key]
    normalized = clean_text(builder)
    if normalized:
        row.add_manual_check(f"builder alias not mapped: {normalized}")
    return normalized


def normalize_zone(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{2}[A-Za-z]", text):
        return text.upper()
    return text


def material_code(raw: Any) -> str | None:
    text = clean_text(raw)
    if not text or text.lower() == "other":
        return None
    lower = text.lower()
    if "colorbond" in lower or "colourbond" in lower:
        suffix = "CB"
    elif any(marker in lower for marker in COLORBOND_COLOUR_MARKERS):
        suffix = "CB"
    elif "stainless" in lower:
        suffix = "SS"
    elif "aluminium" in lower or "aluminum" in lower:
        suffix = "Aluminium"
    elif "galv" in lower or "galvanised" in lower or "galvanized" in lower:
        suffix = "G"
    elif "zinc" in lower:
        suffix = "Z"
    else:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s*mm", text, flags=re.I)
    if not match and suffix == "Aluminium":
        return "Aluminium"
    if not match:
        return None
    thickness = match.group(1)
    if thickness == "1.05":
        thickness = "1"
    else:
        thickness = format(Decimal(thickness).normalize(), "f")
    return f"{thickness}{suffix}"


def cavity_slider_default_material(profile: Any) -> str | None:
    upper = clean_text(profile).upper()
    if "MODERN" in upper:
        return "1.6Z"
    if "DELUXE" in upper:
        return "1Z"
    goods = classify_goods(profile)
    if goods == "MODERN":
        return "1.6Z"
    if goods == "DELUXE":
        return "1Z"
    return None


def join_materials(materials: list[str]) -> str | None:
    seen: list[str] = []
    for item in materials:
        if item and item not in seen:
            seen.append(item)
    if not seen:
        return None
    # Keep thinner material first to match the tracking table style seen in use.
    def sort_key(code: str) -> tuple[float, int, str]:
        match = re.match(r"(\d+(?:\.\d+)?)", code)
        suffix = re.sub(r"^\d+(?:\.\d+)?", "", code)
        suffix_order = {"Z": 0, "G": 1, "CB": 2, "SS": 3, "Aluminium": 4}
        return (float(match.group(1)) if match else 999.0, suffix_order.get(suffix, 99), code)

    return "/".join(sorted(seen, key=sort_key))


def classify_goods(profile: Any, row: ExtractedRow | None = None) -> str | None:
    text = clean_text(profile)
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return None
    key = text.lower()
    if text.upper() in {"WALL", "TYPE"}:
        return None
    if key in PROFILE_ALIASES:
        return PROFILE_ALIASES[key]
    upper = text.upper()
    if "CAVITY" in upper or "SLIDER" in upper:
        return "CS"
    if "SKIN" in upper or "DOOR SKIN" in upper:
        return "DS"
    if "SPLIT" in upper:
        if "DELUXE" in upper or re.search(r"\bDL\b", upper):
            return "SPLIT DL"
        return "SPLIT"
    if "KNOCK" in upper or upper == "KD":
        return "KD"
    if "MODERN" in upper:
        return "MODERN"
    if "DELUXE" in upper:
        return "DELUXE"
    if "CUSTOM" in upper:
        return "COMMERCIAL"
    if "COMMERCIAL" in upper:
        return "COMMERCIAL"
    if any(pattern in key for pattern in RULES.goods_ignore_patterns):
        return None
    if row is not None:
        row.add_manual_check(f"goods type not mapped: {text}")
    return None


def row_context(ws: Any, row_idx: int) -> str:
    return " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 24) + 1))


def classify_goods_with_context(profile: Any, context: str, row: ExtractedRow | None = None) -> str | None:
    text = clean_text(profile)
    upper = text.upper()
    context_upper = context.upper()
    source_upper = (row.source_file if row else "").upper()

    if upper in {"WALL", "TYPE"}:
        return None

    if "DOOR SKIN" in source_upper:
        if "FLAT SHEET" in context_upper:
            return "DS"
        if "CAPPING" in context_upper:
            return None

    if "DOOR STOP BUILD UP" in source_upper:
        return "CP"

    if "CAPPING" in upper:
        return "CAPPING"

    if "FLAT SHEET" in upper:
        return upper

    if "CONCEALED FRAME" in context_upper:
        return "CONCEALED"

    if "/CAV/" in context_upper or "COWDROY" in context_upper or "CLOSING JAMB" in upper:
        return "CS"

    if re.match(r"^(LN|BL)-", upper):
        return "COMMERCIAL"

    return classify_goods(profile, row)


def add_goods(goods_totals: dict[str, float], goods: str | None, qty: Any) -> None:
    qty_num = number_or_none(qty)
    if not goods or qty_num is None or qty_num == 0:
        return
    goods_totals[goods] = goods_totals.get(goods, 0) + float(qty_num)


def profile_is_deluxe_dry_lining(profile: Any) -> bool:
    upper = clean_text(profile).upper()
    return "DRY LINING" in upper or "DR LINING" in upper


def worksheet_goods_bucket(profile: Any, goods: str | None) -> str | None:
    if goods == "DELUXE" and profile_is_deluxe_dry_lining(profile):
        return "DELUXE DRY LINING"
    return goods


def goods_output_type(goods: str) -> str:
    if goods == "DELUXE DRY LINING":
        return "DELUXE"
    return goods


def deluxe_cleats_extra_parts(path: Path, profile: Any, goods: str | None, qty: Any) -> float:
    if "CLEAT" not in path.name.upper() or goods != "DELUXE":
        return 0.0
    qty_num = number_or_none(qty)
    if qty_num is None or qty_num <= 0:
        return 0.0
    profile_upper = clean_text(profile).upper()
    if profile_is_deluxe_dry_lining(profile):
        return float(qty_num) * 6
    return 0.0


def parse_job_number_from_text(value: Any) -> str | int | None:
    text = clean_text(value)
    match = re.search(r"#\s*(\d+[A-Za-z]?)", text)
    if not match:
        match = re.search(r"\b(\d{4,6}[A-Za-z]?)\b", text)
    if not match:
        return None
    raw = match.group(1)
    return int(raw) if raw.isdigit() else raw


def parse_hinge_qty(value: Any) -> float:
    num = number_or_none(value)
    return float(num or 0)


def parse_striker_qty(primary: Any, secondary: Any = None) -> float:
    source = primary if has_value(primary) else secondary
    if not has_value(source):
        return 0.0
    parts = [part for part in re.split(r"\+", clean_text(source)) if has_value(part)]
    return float(len(parts) or 1)


def has_value(value: Any) -> bool:
    text = clean_text(value)
    return bool(text) and text.upper() not in {"-", "NO", "N/A", "NA"}


def write_goods(row: ExtractedRow, goods_totals: dict[str, float]) -> None:
    items = [(goods, qty) for goods, qty in goods_totals.items() if qty]
    if not items:
        return
    items.sort(key=lambda item: (-item[1], item[0]))
    row.values[10] = int(items[0][1]) if items[0][1].is_integer() else items[0][1]
    row.values[11] = goods_output_type(items[0][0])
    if len(items) > 1:
        row.values[12] = int(items[1][1]) if items[1][1].is_integer() else items[1][1]
        row.values[13] = goods_output_type(items[1][0])
    if len(items) > 2:
        row.add_manual_check("more than two goods groups found")


def worksheet_rows(ws: Any) -> list[int]:
    rows: list[int] = []
    for idx in range(10, ws.max_row + 1):
        qty = number_or_none(ws[f"C{idx}"].value)
        profile = clean_text(ws[f"D{idx}"].value)
        material = clean_text(ws[f"A{idx}"].value)
        if qty and (profile or material):
            rows.append(idx)
    return rows


def worksheet_has_standard_detail_header(ws: Any) -> bool:
    return (
        clean_text(ws["A9"].value).upper() == "MATERIAL"
        and clean_text(ws["C9"].value).upper() == "QTY"
        and clean_text(ws["D9"].value).upper() == "PROFILE"
    )


def worksheet_row_is_cavity(ws: Any, row_idx: int) -> bool:
    for idx in range(row_idx, 8, -1):
        first = clean_text(ws[f"A{idx}"].value).upper()
        line = " ".join(clean_text(ws.cell(idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)).upper()
        if idx != row_idx and first == "MATERIAL":
            return False
        if "CAVITY" in line or "SLIDER" in line:
            return True
    return False


def extract_worksheet_workbook(path: Path, row: ExtractedRow, wb: Any, infer_manual: bool = False) -> None:
    ws = wb["Worksheet"]
    data = wb["Data"] if "Data" in wb.sheetnames else None

    row.values[6] = ws["C1"].value or (data["A2"].value if data else None)
    row.values[1] = normalize_po_number(ws["C6"].value or (data["G2"].value if data else None))
    row.values[2] = normalize_builder(ws["C2"].value or (data["C2"].value if data else ""), row)
    row.values[4] = normalize_zone(ws["C7"].value or (data["I2"].value if data else ""))
    row.values[5] = normalize_delivery_address(ws["C4"].value or (data["E2"].value if data else ""))

    delivery = parse_date(ws["C5"].value or (data["F2"].value if data else None))
    row.values[14] = format_date(delivery)
    if infer_manual:
        completion = previous_business_day(delivery)
        row.values[15] = format_date(completion)
        row.values[16] = completion.strftime("%A") if completion else None

    notes = clean_text(ws["I4"].value)
    if notes:
        row.notes.append(f"notes={notes}")

    detail_rows = worksheet_rows(ws)
    materials: list[str] = []
    goods_totals: dict[str, float] = {}
    v_total = 0.0
    x_total = 0.0
    mitre_total = 0.0
    double_qty = 0.0
    over_size_qty = 0.0

    for idx in detail_rows:
        material_raw = ws[f"A{idx}"].value
        qty = number_or_none(ws[f"C{idx}"].value) or 0
        profile = ws[f"D{idx}"].value
        goods = classify_goods_with_context(profile, row_context(ws, idx), row)
        is_cavity = worksheet_row_is_cavity(ws, idx)
        if is_cavity and goods in {None, "MODERN", "DELUXE"}:
            goods = "CS"
        code = material_code(material_raw)
        if code:
            materials.append(code)
        if is_cavity:
            default_code = cavity_slider_default_material(profile)
            if default_code:
                materials.append(default_code)
        if clean_text(material_raw).lower() != "other":
            add_goods(goods_totals, worksheet_goods_bucket(profile, goods), qty)
            max_width = max(width_value(ws[f"G{idx}"].value), width_value(ws[f"I{idx}"].value))
            if qty > 0 and max_width > OVER_SIZE_WIDTH_THRESHOLD:
                over_size_qty += float(qty)

        if not is_cavity:
            hinge = parse_hinge_qty(ws[f"I{idx}"].value)
            striker = parse_striker_qty(ws[f"K{idx}"].value, ws[f"R{idx}"].value)
            sill = 1 if has_value(ws[f"M{idx}"].value) else 0
            line_v = qty * (hinge + striker + sill)
            if goods == "KD":
                line_v += qty * 4
            line_v += deluxe_cleats_extra_parts(path, profile, goods, qty)
            v_total += line_v

        is_double = clean_text(ws[f"O{idx}"].value).upper() == "YES"
        if is_double:
            double_qty += qty

        if goods in {"SPLIT", "SPLIT DL"}:
            mitre_total += qty * (3 if is_double else 2)
        elif goods in {"MODERN", "DELUXE", "COMMERCIAL"}:
            mitre_total += qty
            if is_double:
                mitre_total += 1

        profile_upper = clean_text(profile).upper()
        if "EVOKE KNOCKDOWN" in profile_upper:
            x_total += line_v

    if infer_manual:
        write_over_size(row, workbook_over_size_marker(wb), over_size_qty)

    if infer_manual and row.values[7] is None and ("PALLET" in notes.upper() or double_qty):
        row.add_manual_check("Over Size requires manual entry")

    row.values[9] = join_materials(materials)
    write_goods(row, goods_totals)
    if infer_manual:
        if mitre_total:
            row.values[19] = excel_display_int(mitre_total)
        elif row.values[11] == "KD":
            row.values[19] = 0
        row.values[21] = excel_display_int(v_total)
        if x_total:
            row.values[23] = excel_display_int(x_total)


def extract_nonstandard_worksheet_metadata(path: Path, row: ExtractedRow, wb: Any, infer_manual: bool = False) -> None:
    ws = wb["Worksheet"]

    row.values[6] = (
        parse_job_number_from_text(ws["A1"].value)
        or parse_job_number_from_text(ws["C1"].value)
        or parse_job_number_from_text(path.stem)
    )
    row.values[1] = normalize_po_number(
        sheet1_label_value(ws, {"po", "po no", "po number", "purchase order", "purchase order #"})
    )
    row.values[2] = normalize_builder(sheet1_label_value(ws, {"builder", "invoice"}) or "", row)

    zone, address = extract_sheet1_address(ws)
    if not zone:
        zone = clean_text(sheet1_label_value(ws, {"zone", "location", "delivry zone"})) or None
    row.values[4] = normalize_zone(zone)
    row.values[5] = normalize_delivery_address(address)

    delivery = parse_sheet1_delivery_date(sheet1_label_value(ws, {"delivery date", "delivery d", "date"}), row)
    row.values[14] = format_date(delivery)
    if infer_manual:
        completion = previous_business_day(delivery)
        row.values[15] = format_date(completion)
        row.values[16] = completion.strftime("%A") if completion else None

    row.values[9] = sheet1_material(ws)
    row.add_manual_check("unsupported worksheet detail layout: nonstandard detail header")


def iter_main_detail_rows(ws: Any) -> list[int]:
    rows: list[int] = []
    for idx in range(1, ws.max_row + 1):
        a = ws[f"A{idx}"].value
        b = ws[f"B{idx}"].value
        c = ws[f"C{idx}"].value
        d = ws[f"D{idx}"].value
        f = ws[f"F{idx}"].value
        if any(number_or_none(v) is not None for v in (a, f)) and any(has_value(v) for v in (b, c, d)):
            if clean_text(a).lower() not in {"door #", "qty"} and clean_text(b).lower() != "profile":
                rows.append(idx)
    return rows


def extract_main_workbook(path: Path, row: ExtractedRow, wb: Any, infer_manual: bool = False) -> None:
    ws = wb["Main Sheet"]

    row.values[6] = parse_job_number_from_text(ws["A1"].value)
    raw_po = sheet1_label_value(ws, {"po", "po no", "po number", "purchase order", "purchase order #"}) or ws["B6"].value
    row.values[2] = normalize_builder(sheet1_label_value(ws, {"builder", "invoice"}) or ws["B2"].value or "", row)
    row.values[1] = normalize_danze_door_skin_po_number(raw_po, row.values[2], ws)

    zone, address = extract_sheet1_address(ws)
    if not zone:
        zone = sheet1_label_value(ws, {"delivery zone", "zone", "location", "delivry zone"}) or ws["B7"].value
    if not address:
        address = ws["B4"].value
    row.values[4] = normalize_zone(zone)
    row.values[5] = normalize_delivery_address(address)

    delivery = parse_sheet1_delivery_date(sheet1_label_value(ws, {"delivery date", "delivery d", "date"}) or ws["B5"].value, row)
    row.values[14] = format_date(delivery)
    if infer_manual:
        completion = previous_business_day(delivery)
        row.values[15] = format_date(completion)
        row.values[16] = completion.strftime("%A") if completion else None

    if find_sheet1_profile_headers(ws):
        extract_profile_tables(ws, row, infer_manual=infer_manual)
        return

    materials: list[str] = []
    for idx in range(1, ws.max_row + 1):
        if clean_text(ws[f"A{idx}"].value).lower().startswith("material"):
            code = material_code(f"{clean_text(ws[f'B{idx}'].value)} {clean_text(ws[f'C{idx}'].value)}")
            if code:
                materials.append(code)
    row.values[9] = join_materials(materials)

    goods_totals: dict[str, float] = {}
    v_total = 0.0
    w_total = 0.0
    mitre_total = 0.0
    double_qty = 0.0

    for idx in iter_main_detail_rows(ws):
        qty_a = number_or_none(ws[f"A{idx}"].value)
        qty_f = number_or_none(ws[f"F{idx}"].value)
        profile = ws[f"B{idx}"].value
        goods = classify_goods_with_context(profile, row_context(ws, idx), row)
        hand = clean_text(ws[f"E{idx}"].value).upper()

        # Main Sheet has two shapes:
        # - Danze-style rows: A is a numeric product quantity and F is hinge count.
        # - Commercial rows: B is a profile code and F is the product quantity.
        a_is_real_number = isinstance(ws[f"A{idx}"].value, (int, float))
        if a_is_real_number and qty_a is not None and has_value(profile) and goods != "COMMERCIAL":
            product_qty = float(qty_a)
        elif goods == "COMMERCIAL" and qty_f is not None:
            product_qty = float(qty_f)
        else:
            product_qty = 0.0

        add_goods(goods_totals, goods, product_qty)

        if goods in {"SPLIT", "SPLIT DL"}:
            if hand == "DOUBLE":
                double_qty += product_qty
                mitre_total += product_qty * 3
            else:
                mitre_total += product_qty * 2
        elif goods in {"COMMERCIAL", "MODERN", "DELUXE"}:
            mitre_total += product_qty

        stud = number_or_none(ws[f"K{idx}"].value)
        striker = 1 if has_value(ws[f"H{idx}"].value) else 0
        hinge_qty = number_or_none(ws[f"F{idx}"].value)

        if stud is not None:
            v_total += float(stud) + striker
            if hinge_qty is not None:
                w_total += float(hinge_qty)
        elif hinge_qty is not None:
            v_total += float(hinge_qty)

    if infer_manual:
        write_over_size(row, worksheet_over_size_marker(ws), main_fallback_over_size_qty(ws))
        for idx in range(1, ws.max_row + 1):
            line = " ".join(clean_text(ws.cell(idx, col).value) for col in range(1, ws.max_column + 1))
            if "STILLAGE" in line.upper():
                row.notes.append(f"row {idx}: {line}")
                if row.values[7] is None:
                    row.add_manual_check("Over Size requires manual entry")
                break
        if double_qty and row.values[7] is None:
            row.add_manual_check("Over Size requires manual entry")

    write_goods(row, goods_totals)
    if infer_manual:
        if mitre_total:
            row.values[19] = excel_display_int(mitre_total)
        row.values[21] = excel_display_int(v_total)
        if w_total:
            row.values[22] = excel_display_int(w_total)


def find_label_cell(ws: Any, labels: set[str], max_row: int = 20) -> tuple[int, int] | None:
    for row_idx in range(1, min(ws.max_row, max_row) + 1):
        for col_idx in range(1, ws.max_column + 1):
            label = clean_text(ws.cell(row_idx, col_idx).value).lower().rstrip(":")
            if label in labels:
                return row_idx, col_idx
    return None


def first_right_value(ws: Any, row_idx: int, col_idx: int, max_right: int = 4) -> Any:
    for offset in range(1, max_right + 1):
        value = ws.cell(row_idx, col_idx + offset).value
        if has_value(value):
            return value
    return None


def sheet1_label_value(ws: Any, labels: set[str], max_right: int = 4) -> Any:
    found = find_label_cell(ws, labels)
    if not found:
        return None
    return first_right_value(ws, found[0], found[1], max_right=max_right)


def split_zone_and_address(value: Any) -> tuple[str | None, str | None]:
    text = clean_text(value)
    match = re.match(r"^(\d{2}[A-Za-z])\s*-\s*(.+)$", text)
    if match:
        return match.group(1), match.group(2)
    return None, text or None


def extract_sheet1_address(ws: Any) -> tuple[str | None, str | None]:
    found = find_label_cell(ws, {"delivery address", "delivery a", "address"})
    if not found:
        return None, None
    row_idx, col_idx = found
    values = [
        clean_text(ws.cell(row_idx, col_idx + offset).value)
        for offset in range(1, 5)
        if has_value(ws.cell(row_idx, col_idx + offset).value)
    ]
    if not values:
        return None, None
    if len(values) >= 2 and re.fullmatch(r"\d{2}[A-Za-z]", values[0]):
        return values[0], values[1]
    zone, address = split_zone_and_address(values[0])
    if zone or address:
        return zone, address
    return None, values[0]


def sheet1_material(ws: Any) -> str | None:
    values: list[str] = []
    for row_idx in range(1, min(ws.max_row, 180) + 1):
        for col_idx in range(1, ws.max_column + 1):
            if clean_text(ws.cell(row_idx, col_idx).value).lower().rstrip(":") != "material":
                continue
            material_parts = [
                clean_text(ws.cell(row_idx, col_idx + offset).value)
                for offset in range(1, 5)
                if clean_text(ws.cell(row_idx, col_idx + offset).value)
            ]
            code = material_code(" ".join(material_parts))
            if code:
                values.append(code)
            for offset in range(1, 5):
                code = material_code(ws.cell(row_idx, col_idx + offset).value)
                if code:
                    values.append(code)
    if not values:
        for row_idx in range(1, min(ws.max_row, 15) + 1):
            for col_idx in range(1, ws.max_column + 1):
                code = material_code(ws.cell(row_idx, col_idx).value)
                if code:
                    values.append(code)
    return join_materials(values)


def find_sheet1_profile_header(ws: Any) -> tuple[int, int] | None:
    headers = find_sheet1_profile_headers(ws)
    return headers[0] if headers else None


def find_sheet1_profile_headers(ws: Any) -> list[tuple[int, int]]:
    headers: list[tuple[int, int]] = []
    for row_idx in range(1, min(ws.max_row, 100) + 1):
        for col_idx in range(1, ws.max_column + 1):
            if clean_text(ws.cell(row_idx, col_idx).value).lower() == "profile":
                headers.append((row_idx, col_idx))
    return headers


def sheet1_header_text(ws: Any, header_row: int, col_idx: int) -> str:
    parts = []
    for row_idx in range(max(1, header_row - 2), header_row + 1):
        text = clean_text(ws.cell(row_idx, col_idx).value)
        if text:
            parts.append(text)
    return " ".join(parts).upper()


def sheet1_product_qty_column(ws: Any, header_row: int) -> int | None:
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if not re.search(r"\b(QTY|QUANTITY)\b", header):
            continue
        if any(token in header for token in ("HINGE", "STRIKER", "DYNA", "BOLT", "PLATE")):
            continue
        return col_idx
    return None


def sheet1_product_qty(ws: Any, row_idx: int, profile_col: int, quantity_col: int | None = None) -> float:
    if quantity_col:
        qty = number_or_none(ws.cell(row_idx, quantity_col).value)
        return float(qty or 0)

    left_value = ws.cell(row_idx, profile_col - 1).value if profile_col > 1 else None
    if isinstance(left_value, (int, float)):
        return 1.0
    left_text = clean_text(left_value)
    if has_value(left_value):
        return 1.0
    return 1.0


def sheet1_table_is_cavity(ws: Any, header_row: int) -> bool:
    for row_idx in range(max(1, header_row - 2), header_row + 1):
        line = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        if "CAVITY" in line.upper() or "SLIDER" in line.upper():
            return True
    return False


def sheet1_striker_columns(ws: Any, header_row: int) -> list[int]:
    columns: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if "STRIKER" in header and ("TYPE" in header or "STRIKE" in header):
            columns.append(col_idx)
    return columns


def sheet1_sill_columns(ws: Any, header_row: int) -> list[int]:
    columns: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if re.search(r"\bSILL\b", header):
            columns.append(col_idx)
    return columns


def sheet1_numeric_part_columns(ws: Any, header_row: int) -> list[int]:
    columns: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if "BACKING PLATE" in header or "HINGE" in header or "STRIKER" in header:
            continue
        if any(token in header for token in ("STUD", "DYNA", "2110")):
            columns.append(col_idx)
    return columns


def sheet1_w_part_columns(ws: Any, header_row: int) -> list[int]:
    columns: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if any(token in header for token in ("HINGE HOLDER", "3751", "WS7", "MIB")):
            columns.append(col_idx)
    return columns


def sheet1_hinge_qty_column(ws: Any, header_row: int) -> int | None:
    for col_idx in range(1, ws.max_column + 1):
        header = sheet1_header_text(ws, header_row, col_idx)
        if "HINGE" in header and "QTY" in header:
            return col_idx
    return None


def row_looks_profileless_table_header(ws: Any, row_idx: int) -> bool:
    if row_idx >= ws.max_row:
        return False
    next_first = clean_text(ws.cell(row_idx + 1, 1).value).lower()
    next_second = clean_text(ws.cell(row_idx + 1, 2).value).upper()
    if next_first != "door #" or next_second != "TYPE":
        return False
    line = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)).upper()
    return "PROFILE" not in line and any(token in line for token in ("WALL", "FRAME", "CAVITY", "HINGE"))


def find_sheet1_profileless_table_headers(ws: Any) -> list[tuple[int, int]]:
    headers: list[tuple[int, int]] = []
    for row_idx in range(1, min(ws.max_row, 180) + 1):
        if row_looks_profileless_table_header(ws, row_idx):
            headers.append((row_idx + 1, 2))
    return headers


def sheet1_line_hardware_totals(
    ws: Any,
    row_idx: int,
    header_row: int,
    product_qty: float,
    *,
    include_hardware: bool = True,
) -> tuple[float, float]:
    if not include_hardware:
        return 0.0, 0.0

    part_cols = sheet1_numeric_part_columns(ws, header_row)
    striker_cols = sheet1_striker_columns(ws, header_row)
    sill_cols = sheet1_sill_columns(ws, header_row)
    w_part_cols = sheet1_w_part_columns(ws, header_row)
    hinge_qty_col = sheet1_hinge_qty_column(ws, header_row)

    part_qty = sum(float(number_or_none(ws.cell(row_idx, col).value) or 0) for col in part_cols)
    striker_qty = 1 if any(has_value(ws.cell(row_idx, col).value) for col in striker_cols) else 0
    sill_qty = 1 if any(has_value(ws.cell(row_idx, col).value) for col in sill_cols) else 0
    w_extra_qty = sum(float(number_or_none(ws.cell(row_idx, col).value) or 0) for col in w_part_cols)
    hinge_qty = number_or_none(ws.cell(row_idx, hinge_qty_col).value) if hinge_qty_col else 0

    v_total = product_qty * (part_qty + striker_qty + sill_qty)
    w_total = product_qty * (float(hinge_qty or 0) + w_extra_qty)
    return v_total, w_total


def table_contains_text(ws: Any, start_row: int, end_row: int, token: str) -> bool:
    token_upper = token.upper()
    for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
        line = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)).upper()
        if token_upper in line:
            return True
    return False


def door_skin_capping_only_goods(row: ExtractedRow, context: str, table_has_flat_sheet: bool) -> str | None:
    if "DOOR SKIN" not in row.source_file.upper():
        return None
    if table_has_flat_sheet:
        return None
    if "CAPPING" in context.upper():
        return "DS"
    return None


def extract_profile_tables(ws: Any, row: ExtractedRow, infer_manual: bool = False) -> None:
    materials = (sheet1_material(ws) or "").split("/")
    row.values[9] = join_materials(materials)

    headers = find_sheet1_profile_headers(ws)
    if not headers:
        row.add_manual_check("Sheet1 profile header not found")
        return

    goods_totals: dict[str, float] = {}
    v_total = 0.0
    w_total = 0.0
    x_total = 0.0
    mitre_total = 0.0
    double_qty = 0.0

    for header_index, (header_row, profile_col) in enumerate(headers):
        next_header_row = headers[header_index + 1][0] if header_index + 1 < len(headers) else ws.max_row + 1
        product_qty_col = sheet1_product_qty_column(ws, header_row)
        is_cavity_table = sheet1_table_is_cavity(ws, header_row)
        table_has_flat_sheet = table_contains_text(ws, header_row + 1, next_header_row, "FLAT SHEET")

        for idx in range(header_row + 1, next_header_row):
            if row_looks_profileless_table_header(ws, idx):
                break
            profile = ws.cell(idx, profile_col).value
            if not has_value(profile):
                continue
            if clean_text(ws.cell(idx, 1).value).lower().startswith("material"):
                continue
            if material_code(profile):
                continue
            product_qty = sheet1_product_qty(ws, idx, profile_col, product_qty_col)
            if product_qty <= 0:
                continue

            context = row_context(ws, idx)
            goods = classify_goods_with_context(profile, context, row)
            goods = goods or door_skin_capping_only_goods(row, context, table_has_flat_sheet)
            if is_cavity_table and goods in {None, "MODERN", "DELUXE"}:
                goods = "CS"
            if is_cavity_table:
                default_code = cavity_slider_default_material(profile)
                if default_code:
                    materials.append(default_code)
            add_goods(goods_totals, goods, product_qty)

            hand = clean_text(ws.cell(idx, 5).value).upper()
            is_double = "DOUBLE" in hand
            if not is_double:
                for col_idx in range(1, ws.max_column + 1):
                    if clean_text(ws.cell(idx, col_idx).value).upper() == "YES":
                        header_text = sheet1_header_text(ws, header_row, col_idx)
                        if "DOUBLE" in header_text:
                            is_double = True
                            break
            if is_double:
                double_qty += product_qty

            if goods in {"SPLIT", "SPLIT DL"}:
                mitre_total += product_qty * (3 if is_double else 2)
            elif goods in {"MODERN", "DELUXE", "COMMERCIAL", "KD"}:
                mitre_total += product_qty

            line_v, line_w = sheet1_line_hardware_totals(ws, idx, header_row, product_qty)
            v_total += line_v
            w_total += line_w
            if "EVOKE KNOCKDOWN" in clean_text(profile).upper():
                x_total += line_v

    for header_row, profile_col in find_sheet1_profileless_table_headers(ws):
        product_qty_col = sheet1_product_qty_column(ws, header_row)
        is_cavity_table = sheet1_table_is_cavity(ws, header_row)
        table_has_flat_sheet = table_contains_text(ws, header_row + 1, ws.max_row + 1, "FLAT SHEET")

        for idx in range(header_row + 1, ws.max_row + 1):
            if row_looks_profileless_table_header(ws, idx):
                break
            profile = ws.cell(idx, profile_col).value
            first = clean_text(ws.cell(idx, 1).value)
            if not first or clean_text(ws.cell(idx, 1).value).lower().startswith("material"):
                break
            if not has_value(profile):
                break
            product_qty = sheet1_product_qty(ws, idx, profile_col, product_qty_col)
            if product_qty <= 0:
                continue

            context = row_context(ws, idx)
            goods = classify_goods_with_context(profile, context, row)
            goods = goods or door_skin_capping_only_goods(row, context, table_has_flat_sheet)
            if is_cavity_table and goods in {None, "MODERN", "DELUXE"}:
                goods = "CS"
            if is_cavity_table:
                default_code = cavity_slider_default_material(profile)
                if default_code:
                    materials.append(default_code)
            add_goods(goods_totals, goods, product_qty)

            hand = clean_text(ws.cell(idx, 5).value).upper()
            is_double = "DOUBLE" in hand
            if is_double:
                double_qty += product_qty

            if goods in {"SPLIT", "SPLIT DL"}:
                mitre_total += product_qty * (3 if is_double else 2)
            elif goods in {"MODERN", "DELUXE", "COMMERCIAL", "KD"}:
                mitre_total += product_qty

            line_v, line_w = sheet1_line_hardware_totals(
                ws,
                idx,
                header_row,
                product_qty,
                include_hardware=not is_cavity_table,
            )
            v_total += line_v
            w_total += line_w

    if infer_manual:
        write_over_size(row, worksheet_over_size_marker(ws), profile_tables_over_size_qty(ws))
        for idx in range(1, ws.max_row + 1):
            line = " ".join(clean_text(ws.cell(idx, col).value) for col in range(1, ws.max_column + 1))
            if "STILLAGE" in line.upper() or "PALLET" in line.upper():
                row.notes.append(f"row {idx}: {line}")
                if row.values[7] is None:
                    row.add_manual_check("Over Size requires manual entry")
                break
        if double_qty and row.values[7] is None:
            row.add_manual_check("Over Size requires manual entry")

    row.values[9] = join_materials(materials)
    write_goods(row, goods_totals)
    if infer_manual:
        if mitre_total:
            row.values[19] = excel_display_int(mitre_total)
        elif row.values[11] == "KD":
            row.values[19] = 0
        row.values[21] = excel_display_int(v_total)
        if w_total:
            row.values[22] = excel_display_int(w_total)
        if x_total:
            row.values[23] = excel_display_int(x_total)


def extract_sheet1_workbook(path: Path, row: ExtractedRow, wb: Any, infer_manual: bool = False) -> None:
    ws = wb["Sheet1"]

    row.values[6] = parse_job_number_from_text(ws["A1"].value) or parse_job_number_from_text(path.stem)
    row.values[1] = normalize_po_number(sheet1_label_value(ws, {"po", "order #", "purchase order"}))
    row.values[2] = normalize_builder(
        sheet1_label_value(ws, {"builder", "invoice"}) or "",
        row,
    )

    zone, address = extract_sheet1_address(ws)
    if not zone:
        zone = clean_text(sheet1_label_value(ws, {"zone", "location", "delivry zone"})) or None
    row.values[4] = normalize_zone(zone)
    row.values[5] = normalize_delivery_address(address)

    delivery = parse_sheet1_delivery_date(
        sheet1_label_value(ws, {"delivery date", "delivery d"}),
        row,
    )
    row.values[14] = format_date(delivery)
    if infer_manual:
        completion = previous_business_day(delivery)
        row.values[15] = format_date(completion)
        row.values[16] = completion.strftime("%A") if completion else None

    extract_profile_tables(ws, row, infer_manual=infer_manual)


def compute_parts(row: ExtractedRow) -> None:
    v = number_or_none(row.values[21]) or 0
    w = number_or_none(row.values[22]) or 0
    x = number_or_none(row.values[23]) or 0
    if v or w or x:
        value = float(v) + float(w) * 1.43 + float(x) * 0.43
        row.values[20] = excel_display_int(value)


def extract_workbook(path: Path, infer_manual: bool = False) -> ExtractedRow:
    row = ExtractedRow(source_file=path.name)
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    if "Worksheet" in wb.sheetnames:
        if worksheet_has_standard_detail_header(wb["Worksheet"]):
            extract_worksheet_workbook(path, row, wb, infer_manual=infer_manual)
        else:
            extract_nonstandard_worksheet_metadata(path, row, wb, infer_manual=infer_manual)
    elif "Main Sheet" in wb.sheetnames:
        extract_main_workbook(path, row, wb, infer_manual=infer_manual)
    elif wb.sheetnames == ["Sheet1"]:
        extract_sheet1_workbook(path, row, wb, infer_manual=infer_manual)
    else:
        row.add_manual_check(f"unsupported workbook layout: sheets={wb.sheetnames}")
    if infer_manual:
        compute_parts(row)
    return row


def collect_input_files(input_dir: Path, recursive: bool = False) -> list[Path]:
    paths: list[Path] = []
    iterator = input_dir.rglob("*") if recursive else input_dir.iterdir()
    for path in sorted(iterator):
        if not path.is_file():
            continue
        if path.name.startswith(".~"):
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if path.name.lower().startswith("~$"):
            continue
        if "job track" in path.name.lower():
            continue
        if path.name == "总表.xlsx":
            continue
        paths.append(path)
    return paths


def write_csv(rows: list[ExtractedRow], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(TRACK_HEADERS)
        for row in rows:
            writer.writerow(row.values)


def write_xlsx(rows: list[ExtractedRow], output: Path, input_dir: Path, input_files: list[Path]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单整理"
    ws.append(TRACK_HEADERS)
    for row in rows:
        ws.append(row.values)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (15, 16):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, (dt.date, dt.datetime)):
                cell.number_format = "m/d"

    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 10,
        "E": 10,
        "F": 38,
        "G": 12,
        "H": 12,
        "I": 8,
        "J": 14,
        "K": 10,
        "L": 14,
        "M": 10,
        "N": 14,
        "O": 16,
        "P": 16,
        "Q": 14,
        "R": 10,
        "S": 10,
        "T": 10,
        "U": 10,
        "V": 28,
        "W": 28,
        "X": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    summary = wb.create_sheet("检查摘要")
    unsupported = sum(
        1
        for row in rows
        if any(note.startswith("unsupported workbook layout") for note in row.manual_check)
    )
    manual_check_rows = sum(1 for row in rows if row.manual_check)
    summary_rows = [
        ["项目", "结果"],
        ["输入文件夹", str(input_dir)],
        ["扫描订单 Excel", len(input_files)],
        ["输出订单行", len(rows)],
        ["输出列数", len(TRACK_HEADERS)],
        ["空 Job # 行", sum(1 for row in rows if not normalized_cell(row.values[6]))],
        ["空 PO NUMBER 行", sum(1 for row in rows if not normalized_cell(row.values[1]))],
        ["不支持版式行", unsupported],
        ["需人工检查行", manual_check_rows],
    ]
    for item in summary_rows:
        summary.append(item)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 80
    wb.save(output)


def write_audit_csv(rows: list[ExtractedRow], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "source_file",
            "job",
            "po_number",
            "builder",
            "goods1",
            "notes",
            "manual_check",
            "filled_columns",
            "blank_direct_columns",
        ])
        for row in rows:
            filled = [
                f"{COLUMN_LETTERS[idx]}:{TRACK_HEADERS[idx] or '(blank header)'}"
                for idx, value in enumerate(row.values)
                if value not in (None, "")
            ]
            blank_direct = [
                f"{COLUMN_LETTERS[idx]}:{TRACK_HEADERS[idx] or '(blank header)'}"
                for idx, value in enumerate(row.values)
                if idx not in MANUAL_OR_SCHEDULED_COLUMNS and value in (None, "")
            ]
            writer.writerow([
                row.source_file,
                normalized_cell(row.values[6]),
                normalized_cell(row.values[1]),
                normalized_cell(row.values[2]),
                normalized_cell(row.values[11]),
                " | ".join(row.notes),
                " | ".join(row.manual_check),
                " | ".join(filled),
                " | ".join(blank_direct),
            ])


def normalized_cell(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def normalized_cell_for_compare(value: Any, col_idx: int) -> str:
    text = normalized_cell(value)
    if col_idx == 2 and re.fullmatch(r"\d+", text):
        return text.lstrip("0") or "0"
    if col_idx == 6:
        return text.casefold()
    return text


def address_compare_tokens(value: Any) -> list[str]:
    text = normalized_cell(value).upper().replace("&", " ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    tokens: list[str] = []
    for token in text.split():
        if token == "PICKUP":
            tokens.extend(["PICK", "UP"])
        elif token:
            tokens.append(token)
    return tokens


def contains_compare_sequence(short: list[str], long: list[str]) -> bool:
    if not short or len(short) > len(long):
        return False
    return any(long[idx : idx + len(short)] == short for idx in range(len(long) - len(short) + 1))


def address_cells_match(left: Any, right: Any) -> bool:
    left_tokens = address_compare_tokens(left)
    right_tokens = address_compare_tokens(right)
    if not left_tokens and not right_tokens:
        return True
    if left_tokens == right_tokens:
        return True
    short, long = (left_tokens, right_tokens) if len(left_tokens) <= len(right_tokens) else (right_tokens, left_tokens)
    return len(short) >= 3 and len(long) - len(short) <= 3 and contains_compare_sequence(short, long)


def compare_total(rows: list[ExtractedRow], total_path: Path, diff_path: Path, direct_only: bool = False) -> int:
    wb = openpyxl.load_workbook(total_path, data_only=True)
    ws = wb.active
    actual_by_job: dict[str, tuple[int, list[Any]]] = {}
    for idx in range(2, ws.max_row + 1):
        job = normalized_cell(ws.cell(idx, 7).value)
        if job:
            actual_by_job[job] = (idx, [ws.cell(idx, col).value for col in range(1, 25)])

    diff_path.parent.mkdir(parents=True, exist_ok=True)
    diff_count = 0
    with diff_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["job", "total_row", "column", "header", "extracted", "total", "source_file"])
        for row in rows:
            job = normalized_cell(row.values[6])
            if not job or job not in actual_by_job:
                if job:
                    writer.writerow([job, "", "", "", "", "missing in total", row.source_file])
                    diff_count += 1
                continue
            total_row, actual = actual_by_job[job]
            for idx, (extracted, total) in enumerate(zip(row.values, actual), start=1):
                if direct_only and (idx - 1) in MANUAL_OR_SCHEDULED_COLUMNS:
                    continue
                if idx == 6 and address_cells_match(extracted, total):
                    continue
                if normalized_cell_for_compare(extracted, idx) != normalized_cell_for_compare(total, idx):
                    writer.writerow([
                        job,
                        total_row,
                        COLUMN_LETTERS[idx - 1],
                        TRACK_HEADERS[idx - 1],
                        normalized_cell(extracted),
                        normalized_cell(total),
                        row.source_file,
                    ])
                    diff_count += 1
    return diff_count


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract Ausmet job-tracking rows from order workbooks.")
    parser.add_argument("--input-dir", default=".", help="Directory containing order workbooks.")
    parser.add_argument("--output", default="extracted_job_rows.csv", help="CSV output path.")
    parser.add_argument("--xlsx-output", help="Optional XLSX output path.")
    parser.add_argument("--compare-total", help="Optional total workbook to compare against.")
    parser.add_argument("--diff-output", help="Optional path for comparison diff CSV.")
    parser.add_argument(
        "--diff-direct-only",
        action="store_true",
        help="When comparing, skip manual/scheduled columns H/P/Q/R/S/T/U/V/W/X.",
    )
    parser.add_argument("--audit-output", help="Optional separate audit CSV with source and manual-check notes.")
    parser.add_argument("--recursive", action="store_true", help="Scan input-dir recursively for order workbooks.")
    parser.add_argument("--with-notes", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument(
        "--infer-manual",
        action="store_true",
        help="Best-effort fill for supported manual/scheduled columns such as H/P/Q/T/U/V/W/X; H Over Size uses width > 1024 and packaging notes; T/U/V/W/X use the tracking workbook's 0-decimal display style.",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()

    rows: list[ExtractedRow] = []
    input_files = collect_input_files(input_dir, recursive=args.recursive)
    for path in input_files:
        rows.append(extract_workbook(path, infer_manual=args.infer_manual))

    write_csv(rows, output)
    if args.xlsx_output:
        xlsx_output = Path(args.xlsx_output).expanduser().resolve()
        write_xlsx(rows, xlsx_output, input_dir, input_files)

    if args.audit_output:
        audit_output = Path(args.audit_output).expanduser().resolve()
        write_audit_csv(rows, audit_output)

    if args.compare_total:
        diff_output = (
            Path(args.diff_output).expanduser().resolve()
            if args.diff_output
            else output.with_name(output.stem + "_diff.csv")
        )
        diff_count = compare_total(
            rows,
            Path(args.compare_total).expanduser().resolve(),
            diff_output,
            direct_only=args.diff_direct_only,
        )

    print(f"Scanned {len(input_files)} workbook files")
    print(f"Wrote {len(rows)} rows to {output}")
    if args.xlsx_output:
        print(f"Wrote XLSX output to {xlsx_output}")
    if args.audit_output:
        print(f"Wrote audit report to {audit_output}")
    if args.compare_total:
        print(f"Wrote {diff_count} comparison differences to {diff_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
